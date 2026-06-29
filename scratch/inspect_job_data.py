import openpyxl

job_data_path = "JobData_IR_55800_26-27_20260626_131245.xlsx"
wb = openpyxl.load_workbook(job_data_path, read_only=True)
print("JobData sheetnames:", wb.sheetnames)

if 'GENERAL' in wb.sheetnames:
    ws_gen = wb['GENERAL']
    rows_gen = list(ws_gen.iter_rows(values_only=True))
    headers_gen = [str(h).strip().upper() for h in rows_gen[0]]
    print("\nGENERAL headers:", headers_gen)
    if len(rows_gen) > 1:
        print("GENERAL row 1:")
        for h, val in zip(headers_gen, rows_gen[1]):
            # print non-null values
            if val is not None:
                print(f"  {h}: {val}")

if 'ITEMS' in wb.sheetnames:
    ws_items = wb['ITEMS']
    rows_items = list(ws_items.iter_rows(values_only=True))
    headers_items = [str(h).strip().upper() for h in rows_items[0]]
    print("\nITEMS headers:", headers_items)
    if len(rows_items) > 1:
        print("ITEMS row 1:")
        for h, val in zip(headers_items, rows_items[1]):
            if val is not None:
                print(f"  {h}: {val}")
wb.close()
