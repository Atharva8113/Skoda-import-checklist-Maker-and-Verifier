import openpyxl

wb = openpyxl.load_workbook("JobData_IR_55800_26-27_20260626_131245.xlsx", read_only=True)
if "ITEMS" in wb.sheetnames:
    ws = wb["ITEMS"]
    row1 = next(ws.iter_rows(max_row=1, values_only=True))
    print("ITEMS Headers:", row1)
else:
    print("ITEMS sheet not found")
wb.close()
