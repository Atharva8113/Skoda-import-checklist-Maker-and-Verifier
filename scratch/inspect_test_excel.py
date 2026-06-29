import pandas as pd
import openpyxl

file_path = r"c:\Users\Admin\Documents\NAGARKOT\Documentation\Skoda 1702\Skoda License management system\SAVWIPL Scrip Master TEST.xlsx"

wb = openpyxl.load_workbook(file_path, read_only=True)
print("Sheet names:", wb.sheetnames)
for name in wb.sheetnames:
    print(f"Sheet '{name}' state: {wb[name].sheet_state}")
wb.close()

df = pd.read_excel(file_path, sheet_name="Sheet1")
print("\nSheet1 Columns:")
print(list(df.columns))
print("\nFirst 3 rows:")
print(df.head(3).to_dict(orient='records'))
