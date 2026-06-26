import openpyxl

scrip_path = r"c:\Users\Admin\Documents\NAGARKOT\Documentation\Skoda 1702\Skoda import checklist Maker and Verifier\SAVWIPL Scrip Master 12-06-2026.xlsx"
wb = openpyxl.load_workbook(scrip_path, data_only=False)
ws = wb['Data 14072021']

# Let's inspect formulas for the first few rows
print("=== Scanning Columns for Formulas in Row 2 ===")
for col in range(1, ws.max_column + 1):
    header = ws.cell(row=1, column=col).value
    val = ws.cell(row=2, column=col).value
    if isinstance(val, str) and val.startswith("="):
        print(f"Col {col} ({openpyxl.utils.get_column_letter(col)}) | Header: {header} | Formula: {val}")

wb.close()
