import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import os
import re
import json
import zipfile
import xml.etree.ElementTree as ET
import requests
import pandas as pd
import openpyxl
import threading
from datetime import datetime, timedelta
import socket
import queue
socket.setdefaulttimeout(35)

def safe_float(val):
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        if val != val:
            return 0.0
        return float(val)
    val_str = str(val).strip()
    if not val_str:
        return 0.0
    # Remove commas and spaces
    val_str = val_str.replace(',', '').replace(' ', '')
    try:
        return float(val_str)
    except ValueError:
        return 0.0

def _quick_extract_job_no(filepath):
    """Extract 'Job No' from cell A2 of an xlsx using raw zip+xml streaming.
    
    Bypasses pandas/openpyxl entirely — those hang in Tkinter background
    threads on Windows. This uses only stdlib zipfile + iterparse,
    is fully thread-safe, and stops after finding cell A2.
    Returns the Job No string, or None if not found.
    """
    ns = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    try:
        with zipfile.ZipFile(filepath, 'r') as zf:
            # 1. Load shared strings lookup table
            shared = []
            if 'xl/sharedStrings.xml' in zf.namelist():
                for _, elem in ET.iterparse(zf.open('xl/sharedStrings.xml'), events=('end',)):
                    if elem.tag == f'{{{ns}}}si':
                        parts = [t.text for t in elem.iter(f'{{{ns}}}t') if t.text]
                        shared.append(''.join(parts))
                        elem.clear()
            
            # 2. Find the first worksheet xml file
            sheet_file = None
            for name in sorted(zf.namelist()):
                if name.startswith('xl/worksheets/sheet') and name.endswith('.xml'):
                    sheet_file = name
                    break
            if not sheet_file:
                return None
            
            # 3. Stream-parse worksheet to locate cell A2
            for _, elem in ET.iterparse(zf.open(sheet_file), events=('end',)):
                tag = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag
                if tag == 'c':
                    ref = elem.get('r', '')
                    if ref == 'A2':
                        cell_type = elem.get('t', '')
                        val = ''
                        v_el = elem.find(f'{{{ns}}}v')
                        if v_el is not None and v_el.text:
                            if cell_type == 's':
                                idx = int(v_el.text)
                                val = shared[idx] if idx < len(shared) else ''
                            else:
                                val = v_el.text
                        elem.clear()
                        return val.strip() if val and val.lower() != 'nan' else None
                    elem.clear()
        return None
    except Exception:
        return None

def extract_job_no_from_filename(filepath):
    base = os.path.basename(filepath)
    # Try to find IR_XXXXX_XX-XX
    match = re.search(r'IR_(\d+)_(\d+-\d+)', base, re.IGNORECASE)
    if match:
        return f"IR/{match.group(1)}/{match.group(2)}"
    # Try to find IR_XXXXX
    match2 = re.search(r'IR_([a-zA-Z0-9\-]+)', base, re.IGNORECASE)
    if match2:
        return f"IR/{match2.group(1)}"
    return "Unknown"

def humanize_error(raw_text):
    raw_text = raw_text.strip()
    if not raw_text:
        return "Empty response received from the Google Sheet server."
    if raw_text.startswith("<!DOCTYPE") or "<html" in raw_text.lower():
        title_match = re.search(r"<title>(.*?)</title>", raw_text, re.IGNORECASE | re.DOTALL)
        if title_match:
            title = title_match.group(1).strip()
            return f"Google Sheets Script Error: {title}\n\nPlease check your Google Sheets script configuration or URL."
        return "Google Sheets returned an error page (HTML). This usually means the Google script crashed or the Web App URL is unauthorized/expired."
    if len(raw_text) > 200:
        return raw_text[:200] + "..."
    return raw_text

# --- MONKEYPATCH FOR OPENPYXL TO WRITE FORMULA CACHED VALUES ---
from xml.etree.ElementTree import Element, SubElement
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula
from openpyxl.cell.rich_text import CellRichText
from openpyxl.cell._writer import _set_attributes, safe_string, whitespace

# Registry for cached values of formulas (mapped by (sheet_title, cell_coordinate))
CACHED_VALUES = {}

def custom_etree_write_cell(xf, worksheet, cell, styled=None):
    value, attributes = _set_attributes(cell, styled)

    el = Element("c", attributes)
    if value is None or value == "":
        xf.write(el)
        return

    is_formula = (cell.data_type == 'f')

    if is_formula:
        attrib = {}
        if isinstance(value, ArrayFormula):
            attrib = dict(value)
            value = value.text
        elif isinstance(value, DataTableFormula):
            attrib = dict(value)
            value = None

        formula = SubElement(el, 'f', attrib)
        if value is not None and not attrib.get('t') == "dataTable":
            formula.text = value[1:]
            
            # Check if there is a custom cached value
            key = (worksheet.title, cell.coordinate)
            if key in CACHED_VALUES:
                value = CACHED_VALUES[key]
            else:
                value = None

    if cell.data_type == 's':
        if isinstance(value, CellRichText):
            el.append(value.to_tree())
        else:
            inline_string = Element("is")
            text = Element('t')
            text.text = value
            whitespace(text)
            inline_string.append(text)
            el.append(inline_string)
    else:
        cell_content = SubElement(el, 'v')
        if value is not None:
            cell_content.text = safe_string(value)

    xf.write(el)

# Apply monkeypatch
import openpyxl.cell._writer
import openpyxl.worksheet._writer
openpyxl.cell._writer.etree_write_cell = custom_etree_write_cell
openpyxl.worksheet._writer.write_cell = custom_etree_write_cell
# -------------------------------------------------------------

# Design Tokens (Corporate Blue & White Theme)
COLOR_PRIMARY = "#0F3057"       # Corporate Dark Blue
COLOR_SECONDARY = "#00587A"     # Medium Blue
COLOR_BG = "#F4F6F9"            # Light Gray-Blue Background
COLOR_CARD = "#FFFFFF"          # White Card Background
COLOR_TEXT = "#333333"          # Dark Gray Text
COLOR_ACCENT = "#17B978"        # Green accent for success / action
COLOR_ERROR = "#FF4D4D"         # Red accent for error
FONT_FAMILY = "Segoe UI"

CONFIG_FILE = "config.json"

def load_config():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r') as f:
                return json.load(f)
        except Exception:
            pass
    return {"google_sheet_url": "", "security_token": "NagarkotSkoda2026"}

def save_config(config):
    try:
        with open(CONFIG_FILE, 'w') as f:
            json.dump(config, f, indent=4)
    except Exception:
        pass

class LicenseAutomationApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Nagarkot Skoda License Automation Tool")
        self.root.geometry("1100x850")
        self.root.configure(bg=COLOR_BG)
        
        # Configuration
        self.config = load_config()
        self.google_sheet_url = tk.StringVar(value=self.config.get("google_sheet_url", ""))
        self.security_token = tk.StringVar(value=self.config.get("security_token", "NagarkotSkoda2026"))
        
        # State Variables
        self.job_data_path = tk.StringVar()
        self.item_report_path = tk.StringVar()
        self.new_lic_excel_path = tk.StringVar()
        self.rectify_jd_path = tk.StringVar()
        self.blacklist_jd_path = tk.StringVar()
        self.blacklist_licenses = []
        
        # Blacklist Tab Reallocation State
        self.blacklist_active_lics = []
        self.blacklist_selected_lic_nos = set()
        self.blacklist_selected_to_blacklist = set()
        self.blacklist_parsed_debits = []
        self.blacklist_search_var = tk.StringVar()
        self.blacklist_search_var.trace_add("write", lambda *args: self.populate_blacklist_select_tree())
        self.blacklist_affected_items = []
        self.blacklist_job_info = {}
        self.blacklist_reallocations_preview = []
        
        self.job_info = {}
        self.active_licenses = []
        self.used_licenses = []
        self.required_duty = 0.0
        self.selected_duty_covered = 0.0
        self.estimated_debit = 0.0
        self.item_duties = []
        
        self.gui_queue = queue.Queue()
        self.root.after(100, self.process_gui_queue)
        
        self.create_styles()
        self.build_ui()
        
        # Initial database view load if configured
        if self.google_sheet_url.get().strip():
            self.refresh_db_view()
            self.refresh_blacklist()

    def create_styles(self):
        """Configure custom styles for ttk widgets."""
        self.style = ttk.Style()
        self.style.theme_use('clam')
        
        # Frame styles
        self.style.configure('TFrame', background=COLOR_BG)
        self.style.configure('Card.TFrame', background=COLOR_CARD, relief='flat', borderwidth=0)
        
        # Label styles
        self.style.configure('TLabel', background=COLOR_BG, foreground=COLOR_TEXT, font=(FONT_FAMILY, 10))
        self.style.configure('Header.TLabel', background=COLOR_PRIMARY, foreground='#FFFFFF', font=(FONT_FAMILY, 16, 'bold'))
        self.style.configure('Section.TLabel', background=COLOR_CARD, foreground=COLOR_PRIMARY, font=(FONT_FAMILY, 12, 'bold'))
        self.style.configure('Summary.TLabel', background=COLOR_CARD, foreground=COLOR_TEXT, font=(FONT_FAMILY, 10, 'bold'))
        
        # Button styles
        self.style.configure('TButton', font=(FONT_FAMILY, 10, 'bold'), borderwidth=0)
        self.style.map('TButton',
            background=[('active', COLOR_SECONDARY), ('!disabled', COLOR_PRIMARY)],
            foreground=[('active', '#FFFFFF'), ('!disabled', '#FFFFFF')]
        )
        
        self.style.configure('Action.TButton', font=(FONT_FAMILY, 11, 'bold'))
        self.style.map('Action.TButton',
            background=[('active', '#149F67'), ('!disabled', COLOR_ACCENT)],
            foreground=[('active', '#FFFFFF'), ('!disabled', '#FFFFFF')]
        )

        # Entry styles
        self.style.configure('TEntry', fieldbackground='#FFFFFF', bordercolor='#CCCCCC', lightcolor='#CCCCCC', darkcolor='#CCCCCC')
        
        # Treeview styles
        self.style.configure('Treeview',
            background=COLOR_CARD,
            fieldbackground=COLOR_CARD,
            foreground=COLOR_TEXT,
            rowheight=25,
            font=(FONT_FAMILY, 10)
        )
        self.style.configure('Treeview.Heading',
            background=COLOR_PRIMARY,
            foreground='#FFFFFF',
            font=(FONT_FAMILY, 10, 'bold'),
            borderwidth=1,
            relief='flat'
        )
        self.style.map('Treeview.Heading',
            background=[('active', COLOR_SECONDARY), ('!disabled', COLOR_PRIMARY)],
            foreground=[('active', '#FFFFFF'), ('!disabled', '#FFFFFF')]
        )

    def build_ui(self):
        """Assemble the main window components."""
        # --- HEADER BLOCK ---
        header_frame = tk.Frame(self.root, bg='#FFFFFF', height=60)
        header_frame.pack(fill='x', side='top')
        header_frame.pack_propagate(False)
        
        # Logo handling (Pillow LANCZOS high-quality resampling for crisp rendering)
        logo_loaded = False
        logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)) if '__file__' in globals() else os.getcwd(), "logo.png")
        if os.path.exists(logo_path):
            try:
                from PIL import Image, ImageTk
                pil_img = Image.open(logo_path)
                target_h = 36
                aspect = pil_img.width / pil_img.height
                target_w = int(target_h * aspect)
                resample_mode = getattr(Image, 'Resampling', Image).LANCZOS
                resized_pil = pil_img.resize((target_w, target_h), resample_mode)
                self.logo_img = ImageTk.PhotoImage(resized_pil)
                
                logo_label = tk.Label(header_frame, image=self.logo_img, bg='#FFFFFF')
                logo_label.place(x=20, rely=0.5, anchor='w')
                logo_loaded = True
            except Exception as e:
                try:
                    self.logo_img = tk.PhotoImage(file=logo_path)
                    sub_factor = max(1, self.logo_img.height() // 36)
                    self.logo_img = self.logo_img.subsample(sub_factor)
                    logo_label = tk.Label(header_frame, image=self.logo_img, bg='#FFFFFF')
                    logo_label.place(x=20, rely=0.5, anchor='w')
                    logo_loaded = True
                except Exception:
                    pass
        
        if not logo_loaded:
            logo_label = tk.Label(header_frame, text="NAGARKOT", fg=COLOR_PRIMARY, bg='#FFFFFF', font=(FONT_FAMILY, 14, 'bold'))
            logo_label.place(x=20, rely=0.5, anchor='w')
            
        # Centrally aligned Title across header width
        title_label = tk.Label(
            header_frame, 
            text="SKODA LICENSE MANAGER", 
            fg=COLOR_PRIMARY, 
            bg='#FFFFFF', 
            font=(FONT_FAMILY, 18, 'bold')
        )
        title_label.place(relx=0.5, rely=0.5, anchor='center')
        
        # Subtle divider bar under header
        accent_line = tk.Frame(self.root, bg=COLOR_PRIMARY, height=3)
        accent_line.pack(fill='x', side='top')
        
        # --- NOTEBOOK FOR TABS ---
        notebook = ttk.Notebook(self.root)
        notebook.pack(fill='both', expand=True, padx=15, pady=5)
        
        # Tab 1: License Automation
        tab_automation = ttk.Frame(notebook)
        notebook.add(tab_automation, text="License Automation")
        # Tab 2: Database Master
        tab_db = ttk.Frame(notebook)
        notebook.add(tab_db, text="Database Master")
        
        # Tab 3: Blacklist License
        tab_blacklist = ttk.Frame(notebook)
        notebook.add(tab_blacklist, text="Blacklist License")
        self.build_blacklist_tab(tab_blacklist)
        
        # Tab 4: Rectify License
        tab_rectify = ttk.Frame(notebook)
        notebook.add(tab_rectify, text="Rectify License")
        self.build_rectify_tab(tab_rectify)
        
        # --- TAB 1 CONTENT (LICENSE AUTOMATION) ---
        main_frame = ttk.Frame(tab_automation, padding="10")
        main_frame.pack(fill='both', expand=True)
        
        # Left Panel (Inputs and Job Details)
        left_panel = ttk.Frame(main_frame)
        left_panel.pack(side='left', fill='both', expand=True, padx=(0, 10))
        
        # Card 1: File Pickers
        files_card = ttk.Frame(left_panel, style='Card.TFrame', padding="15")
        files_card.pack(fill='x', side='top', pady=(0, 10))
        
        ttk.Label(files_card, text="File Configuration", style='Section.TLabel').pack(anchor='w', pady=(0, 10))
        
        # JobData Path Input
        row_jd = ttk.Frame(files_card, style='Card.TFrame')
        row_jd.pack(fill='x', pady=5)
        ttk.Label(row_jd, text="JobData Excel:", width=18, anchor='w', style='Summary.TLabel').pack(side='left')
        ttk.Entry(row_jd, textvariable=self.job_data_path).pack(side='left', fill='x', expand=True, padx=5)
        ttk.Button(row_jd, text="Browse", command=self.browse_job_data).pack(side='left')
        
        # Item Report Path Input
        row_ir = ttk.Frame(files_card, style='Card.TFrame')
        row_ir.pack(fill='x', pady=5)
        ttk.Label(row_ir, text="Item Report Excel:", width=18, anchor='w', style='Summary.TLabel').pack(side='left')
        ttk.Entry(row_ir, textvariable=self.item_report_path).pack(side='left', fill='x', expand=True, padx=5)
        ttk.Button(row_ir, text="Browse", command=self.browse_item_report).pack(side='left')
        

        
        # Load Data Button
        self.load_btn = ttk.Button(files_card, text="LOAD AND ANALYZE FILES", command=self.load_and_analyze_data)
        self.load_btn.pack(fill='x', pady=(15, 0))
        
        # Card 2: Job details summary
        self.summary_card = ttk.Frame(left_panel, style='Card.TFrame', padding="15")
        self.summary_card.pack(fill='both', expand=True)
        
        ttk.Label(self.summary_card, text="Job Details Summary", style='Section.TLabel').pack(anchor='w', pady=(0, 10))
        
        self.summary_text_frame = ttk.Frame(self.summary_card, style='Card.TFrame')
        self.summary_text_frame.pack(fill='both', expand=True)
        
        # Initial Placeholder text in Summary
        self.summary_placeholder = ttk.Label(self.summary_text_frame, text="Please load the JobData and Item Report files to view summary.", foreground="#777777", style='TLabel')
        self.summary_placeholder.pack(anchor='w', pady=10)
        
        # Right Panel (Licenses and Logging)
        right_panel = ttk.Frame(main_frame)
        right_panel.pack(side='right', fill='both', expand=True, padx=(10, 0))
        
        # Card 3: Active Licenses Table
        lic_card = ttk.Frame(right_panel, style='Card.TFrame', padding="15")
        lic_card.pack(fill='both', expand=True, side='top', pady=(0, 10))
        
        ttk.Label(lic_card, text="Active Licenses (Fetched from Cloud)", style='Section.TLabel').pack(anchor='w', pady=(0, 5))
        ttk.Label(lic_card, text="Retains database master order. Check/uncheck to select licenses for allocation.", font=(FONT_FAMILY, 9, 'italic'), foreground="#666666").pack(anchor='w', pady=(0, 10))
        
        # Treeview for active licenses
        tree_scroll_y = ttk.Scrollbar(lic_card, orient='vertical')
        tree_scroll_y.pack(side='right', fill='y')
        tree_scroll_x = ttk.Scrollbar(lic_card, orient='horizontal')
        tree_scroll_x.pack(side='bottom', fill='x')
        
        self.lic_tree = ttk.Treeview(
            lic_card, 
            columns=('selected', 'lic_no', 'port', 'type', 'balance', 'expiry'), 
            show='headings', 
            yscrollcommand=tree_scroll_y.set,
            xscrollcommand=tree_scroll_x.set,
            selectmode='none'
        )
        self.lic_tree.pack(fill='both', expand=True)
        tree_scroll_y.config(command=self.lic_tree.yview)
        tree_scroll_x.config(command=self.lic_tree.xview)
        
        self.lic_tree.tag_configure('near_expiry', foreground=COLOR_ERROR)
        self.lic_tree.tag_configure('checked', background="#E8F5E9")
        
        # Column Headings
        self.lic_tree.heading('selected', text="[Select]")
        self.lic_tree.heading('lic_no', text="License No.")
        self.lic_tree.heading('port', text="Port")
        self.lic_tree.heading('type', text="Type")
        self.lic_tree.heading('balance', text="Balance (INR)")
        self.lic_tree.heading('expiry', text="Expiry Date")
        
        # Column Widths
        self.lic_tree.column('selected', width=80, anchor='center')
        self.lic_tree.column('lic_no', width=130, anchor='center')
        self.lic_tree.column('port', width=80, anchor='center')
        self.lic_tree.column('type', width=90, anchor='center')
        self.lic_tree.column('balance', width=130, anchor='e')
        self.lic_tree.column('expiry', width=110, anchor='center')
        
        self.lic_tree.bind("<ButtonRelease-1>", self.on_tree_click)
        
        # Card 4: Run Actions and Live log
        run_card = ttk.Frame(right_panel, style='Card.TFrame', padding="15")
        run_card.pack(fill='x', side='bottom')
        
        self.run_btn = ttk.Button(run_card, text="GENERATE logisys EXCEL & DEBIT LICENSES", style='Action.TButton', command=self.run_license_automation, state='disabled')
        self.run_btn.pack(fill='x', pady=(0, 10))
        
        # Log view
        ttk.Label(run_card, text="Execution Logs", style='Section.TLabel').pack(anchor='w', pady=(0, 5))
        log_scroll = ttk.Scrollbar(run_card)
        log_scroll.pack(side='right', fill='y')
        
        self.log_txt = tk.Text(run_card, height=8, wrap='word', yscrollcommand=log_scroll.set, font=('Consolas', 9), bg='#FAF9F6', fg='#333333', borderwidth=1, relief='solid')
        self.log_txt.pack(fill='x')
        log_scroll.config(command=self.log_txt.yview)
        self.log("System initialized. Configure Google Sheets URL, load data to begin.")
        
        # --- TAB 2 CONTENT (DATABASE MASTER) ---
        db_main_frame = ttk.Frame(tab_db, padding="10")
        db_main_frame.pack(fill='both', expand=True)
        
        # Top Card: Configuration
        config_card = ttk.Frame(db_main_frame, style='Card.TFrame', padding="10")
        config_card.pack(fill='x', side='top', pady=(0, 10))
        
        ttk.Label(config_card, text="Google Sheets Database Configuration", style='Section.TLabel').pack(anchor='w', pady=(0, 5))
        
        row_url = ttk.Frame(config_card, style='Card.TFrame')
        row_url.pack(fill='x', pady=2)
        ttk.Label(row_url, text="Google Web App URL:", width=22, anchor='w', style='Summary.TLabel').pack(side='left')
        ttk.Entry(row_url, textvariable=self.google_sheet_url).pack(side='left', fill='x', expand=True, padx=5)
        ttk.Button(row_url, text="Save Configuration", command=self.save_app_config).pack(side='left')
        
        # Middle Card: Spreadsheet View
        sheet_card = ttk.Frame(db_main_frame, style='Card.TFrame', padding="15")
        sheet_card.pack(fill='both', expand=True, side='top', pady=(0, 10))
        
        row_sheet_hdr = ttk.Frame(sheet_card, style='Card.TFrame')
        row_sheet_hdr.pack(fill='x', pady=(0, 5))
        ttk.Label(row_sheet_hdr, text="Spreadsheet View (Licenses in Cloud)", style='Section.TLabel').pack(side='left')
        ttk.Button(row_sheet_hdr, text="Refresh Database View", command=self.refresh_db_view).pack(side='right')
        
        # Treeview for viewing all columns in Data sheet
        sheet_scroll_y = ttk.Scrollbar(sheet_card, orient='vertical')
        sheet_scroll_y.pack(side='right', fill='y')
        sheet_scroll_x = ttk.Scrollbar(sheet_card, orient='horizontal')
        sheet_scroll_x.pack(side='bottom', fill='x')
        
        self.db_tree = ttk.Treeview(
            sheet_card,
            columns=('lic_no', 'lic_date', 'port', 'type', 'value', 'expiry', 'no_boe', 'used_val', 'balance', 'job_no'),
            show='headings',
            yscrollcommand=sheet_scroll_y.set,
            xscrollcommand=sheet_scroll_x.set
        )
        self.db_tree.pack(fill='both', expand=True)
        sheet_scroll_y.config(command=self.db_tree.yview)
        sheet_scroll_x.config(command=self.db_tree.xview)
        
        # DB View Columns Headers
        self.db_tree.heading('lic_no', text="License No")
        self.db_tree.heading('lic_date', text="Lic Date")
        self.db_tree.heading('port', text="Port")
        self.db_tree.heading('type', text="Type")
        self.db_tree.heading('value', text="Value (INR)")
        self.db_tree.heading('expiry', text="Expiry Date")
        self.db_tree.heading('no_boe', text="Count of Job")
        self.db_tree.heading('used_val', text="Used Value (INR)")
        self.db_tree.heading('balance', text="Balance (INR)")
        self.db_tree.heading('job_no', text="Job List")
        
        # Column width / align
        self.db_tree.column('lic_no', width=120, anchor='center')
        self.db_tree.column('lic_date', width=90, anchor='center')
        self.db_tree.column('port', width=80, anchor='center')
        self.db_tree.column('type', width=90, anchor='center')
        self.db_tree.column('value', width=110, anchor='e')
        self.db_tree.column('expiry', width=100, anchor='center')
        self.db_tree.column('no_boe', width=80, anchor='center')
        self.db_tree.column('used_val', width=120, anchor='e')
        self.db_tree.column('balance', width=120, anchor='e')
        self.db_tree.column('job_no', width=200, anchor='w')
        
        # Bottom Card: Add Licenses Options (Side by side)
        add_card = ttk.Frame(db_main_frame, style='Card.TFrame', padding="10")
        add_card.pack(fill='x', side='bottom')
        
        ttk.Label(add_card, text="Add New Licenses to Cloud", style='Section.TLabel').pack(anchor='w', pady=(0, 10))
        
        # Left half: Paste Text
        paste_frame = ttk.Frame(add_card, style='Card.TFrame')
        paste_frame.pack(side='left', fill='both', expand=True, padx=(0, 10))
        
        ttk.Label(paste_frame, text="Method A: Paste License Rows (Tab or Comma separated)\nOrder: LICNO/ | LIC Date | Port | Type | Value | [Expiry Date]", justify='left', style='Summary.TLabel').pack(anchor='w', pady=(0, 5))
        
        self.paste_txt = tk.Text(paste_frame, height=5, wrap='none', font=('Consolas', 9), bg='#FFFFFF', fg='#333333', borderwidth=1, relief='solid')
        self.paste_txt.pack(fill='both', expand=True, pady=5)
        
        self.paste_btn = ttk.Button(paste_frame, text="Push Pasted Licenses to Cloud", command=self.push_pasted_licenses)
        self.paste_btn.pack(fill='x')
        
        # Right half: Excel upload
        excel_frame = ttk.Frame(add_card, style='Card.TFrame')
        excel_frame.pack(side='right', fill='both', expand=True, padx=(10, 0))
        
        ttk.Label(excel_frame, text="Method B: Select Excel file with license sheets", style='Summary.TLabel').pack(anchor='w', pady=(0, 5))
        
        row_pick = ttk.Frame(excel_frame, style='Card.TFrame')
        row_pick.pack(fill='x', pady=5)
        ttk.Entry(row_pick, textvariable=self.new_lic_excel_path).pack(side='left', fill='x', expand=True, padx=(0, 5))
        ttk.Button(row_pick, text="Browse", command=self.browse_new_lic_excel).pack(side='left')
        
        ttk.Label(excel_frame, text="Supported headers: LICNO/, LIC Date, Port, Type, Value, Expiry Date\nEnsure headers or columns align correctly.", font=(FONT_FAMILY, 8, 'italic'), foreground="#777777").pack(anchor='w', pady=5)
        
        self.excel_btn = ttk.Button(excel_frame, text="Import and Push Excel to Cloud", command=self.push_excel_licenses)
        self.excel_btn.pack(fill='x')
        
        # --- FOOTER BLOCK ---
        footer_frame = tk.Frame(self.root, bg="#E2E2E2", height=36)
        footer_frame.pack(fill='x', side='bottom')
        footer_frame.pack_propagate(False)
        
        # Left side container for logo & company name
        footer_left = tk.Frame(footer_frame, bg="#E2E2E2")
        footer_left.pack(side='left', padx=15, fill='y')
        
        # Logo handling for footer (High-quality Pillow resampling)
        if os.path.exists(logo_path):
            try:
                from PIL import Image, ImageTk
                pil_img_ft = Image.open(logo_path)
                target_h_ft = 20
                aspect_ft = pil_img_ft.width / pil_img_ft.height
                target_w_ft = int(target_h_ft * aspect_ft)
                resample_mode_ft = getattr(Image, 'Resampling', Image).LANCZOS
                resized_pil_ft = pil_img_ft.resize((target_w_ft, target_h_ft), resample_mode_ft)
                self.footer_logo_img = ImageTk.PhotoImage(resized_pil_ft)
                
                footer_logo_lbl = tk.Label(footer_left, image=self.footer_logo_img, bg="#E2E2E2")
                footer_logo_lbl.pack(side='left', padx=(0, 8), pady=4)
            except Exception:
                try:
                    self.footer_logo_img = tk.PhotoImage(file=logo_path)
                    sub_factor = max(1, self.footer_logo_img.height() // 20)
                    self.footer_logo_img = self.footer_logo_img.subsample(sub_factor)
                    footer_logo_lbl = tk.Label(footer_left, image=self.footer_logo_img, bg="#E2E2E2")
                    footer_logo_lbl.pack(side='left', padx=(0, 8), pady=4)
                except Exception:
                    pass
                
        footer_brand_lbl = tk.Label(
            footer_left, 
            text="Nagarkot Forwarders Pvt. Ltd. ©", 
            bg="#E2E2E2", 
            fg="#333333", 
            font=(FONT_FAMILY, 9, 'bold')
        )
        footer_brand_lbl.pack(side='left', pady=7)
        
        # Right side container for system info
        footer_right = tk.Label(
            footer_frame, 
            text="Developed for Skoda Imports Automation | Central Google Sheet Mode", 
            bg="#E2E2E2", 
            fg="#666666", 
            font=(FONT_FAMILY, 8, 'italic')
        )
        footer_right.pack(side='right', padx=15, pady=8)


    def build_blacklist_tab(self, parent):
        """Build the layout and widgets for Tab 3: Blacklist License."""
        main_pane = ttk.PanedWindow(parent, orient='horizontal')
        main_pane.pack(fill='both', expand=True, padx=10, pady=10)
        
        # --- LEFT PANEL ---
        left_frame = ttk.Frame(main_pane, style='Card.TFrame', padding=15)
        main_pane.add(left_frame, weight=1)
        
        ttk.Label(left_frame, text="Blacklist License & Re-allocate", style='Section.TLabel').pack(anchor='w', pady=(0, 15))
        
        # File Picker
        row_file = ttk.Frame(left_frame, style='Card.TFrame')
        row_file.pack(fill='x', pady=5)
        ttk.Label(row_file, text="Processed JobData:", width=18, anchor='w', style='Summary.TLabel').pack(side='left')
        ttk.Entry(row_file, textvariable=self.blacklist_jd_path).pack(side='left', fill='x', expand=True, padx=5)
        ttk.Button(row_file, text="Browse", command=self.browse_blacklist_jd).pack(side='left')
        
        # Parse Button
        self.blacklist_parse_btn = ttk.Button(left_frame, text="PARSE DEBITED LICENSES", command=self.parse_blacklist_jd)
        self.blacklist_parse_btn.pack(fill='x', pady=10)
        
        # Scrollable Treeview for selecting multiple licenses to blacklist
        row_lic = ttk.Frame(left_frame, style='Card.TFrame')
        row_lic.pack(fill='both', expand=True, pady=5)
        ttk.Label(row_lic, text="Select Licenses to Blacklist:", style='Summary.TLabel').pack(anchor='w', pady=(0, 5))
        
        # Searchbar and Dropdown Checker Frame
        search_filter_frame = ttk.Frame(row_lic, style='Card.TFrame')
        search_filter_frame.pack(fill='x', pady=(0, 5))
        
        ttk.Label(search_filter_frame, text="Search:", style='Summary.TLabel').pack(side='left', padx=(0, 5))
        search_entry = ttk.Entry(search_filter_frame, textvariable=self.blacklist_search_var, width=15)
        search_entry.pack(side='left', fill='x', expand=True, padx=(0, 10))
        
        self.blacklist_dropdown_checker = ttk.Combobox(
            search_filter_frame, 
            values=["Check All", "Uncheck All"], 
            state='readonly', 
            width=15
        )
        self.blacklist_dropdown_checker.set("[Select All/None]")
        self.blacklist_dropdown_checker.pack(side='right', padx=(5, 0))
        self.blacklist_dropdown_checker.bind("<<ComboboxSelected>>", self.on_blacklist_dropdown_checker_changed)
        
        tree_frame = ttk.Frame(row_lic)
        tree_frame.pack(fill='both', expand=True)
        
        self.blacklist_select_tree = ttk.Treeview(
            tree_frame, 
            columns=('Select', 'LicenseNo', 'Type', 'Port'), 
            show='headings', 
            height=3
        )
        self.blacklist_select_tree.heading('Select', text='[Select]')
        self.blacklist_select_tree.heading('LicenseNo', text='License No.')
        self.blacklist_select_tree.heading('Type', text='Type')
        self.blacklist_select_tree.heading('Port', text='Port')
        
        self.blacklist_select_tree.column('Select', width=65, anchor='center')
        self.blacklist_select_tree.column('LicenseNo', width=130, anchor='center')
        self.blacklist_select_tree.column('Type', width=90, anchor='center')
        self.blacklist_select_tree.column('Port', width=90, anchor='center')
        
        scroll = ttk.Scrollbar(tree_frame, orient='vertical', command=self.blacklist_select_tree.yview)
        self.blacklist_select_tree.configure(yscrollcommand=scroll.set)
        
        self.blacklist_select_tree.pack(side='left', fill='both', expand=True)
        scroll.pack(side='right', fill='y')
        
        self.blacklist_select_tree.bind('<ButtonRelease-1>', self.on_blacklist_select_tree_click)
        self.blacklist_select_tree.tag_configure('checked', background="#E8F5E9")
        
        # Reason Entry
        row_reason = ttk.Frame(left_frame, style='Card.TFrame')
        row_reason.pack(fill='x', pady=10)
        ttk.Label(row_reason, text="Reason for Blacklisting:", style='Summary.TLabel').pack(anchor='w', pady=(0, 5))
        self.blacklist_reason_entry = ttk.Entry(row_reason)
        self.blacklist_reason_entry.pack(fill='x')
        
        # Job Details Summary Preview
        self.blacklist_preview_card = ttk.LabelFrame(left_frame, text="Job Details Summary Preview", padding=10)
        self.blacklist_preview_card.pack(fill='both', expand=True, pady=10)
        
        self.blacklist_summary_lbl = ttk.Label(
            self.blacklist_preview_card, 
            text="Please upload processed JobData and parse to view preview.", 
            foreground="#666666", 
            justify='left', 
            anchor='nw', 
            font=(FONT_FAMILY, 10)
        )
        self.blacklist_summary_lbl.pack(fill='both', expand=True)
        
        # Action Button
        self.blacklist_execute_btn = ttk.Button(
            left_frame, 
            text="BLACKLIST & RE-ALLOCATE DEBITS", 
            command=self.run_blacklist_reallocation, 
            state='disabled', 
            style='Action.TButton'
        )
        self.blacklist_execute_btn.pack(fill='x', pady=(10, 0))
        
        # --- RIGHT PANEL ---
        right_frame = ttk.Frame(main_pane, style='Card.TFrame', padding=15)
        main_pane.add(right_frame, weight=2)
        
        # Card 1: Active Licenses Table for Reallocation
        ttk.Label(right_frame, text="Active Licenses (Eligible for Re-allocation)", style='Section.TLabel').pack(anchor='w', pady=(0, 5))
        ttk.Label(right_frame, text="Check/uncheck to select licenses to be used for reallocation.", font=(FONT_FAMILY, 9, 'italic'), foreground="#666666").pack(anchor='w', pady=(0, 10))
        
        self.blacklist_lic_tree_frame = ttk.Frame(right_frame)
        self.blacklist_lic_tree_frame.pack(fill='both', expand=True, pady=(0, 10))
        
        scroll_y = ttk.Scrollbar(self.blacklist_lic_tree_frame, orient='vertical')
        scroll_y.pack(side='right', fill='y')
        scroll_x = ttk.Scrollbar(self.blacklist_lic_tree_frame, orient='horizontal')
        scroll_x.pack(side='bottom', fill='x')
        
        self.blacklist_lic_tree = ttk.Treeview(
            self.blacklist_lic_tree_frame,
            columns=('selected', 'lic_no', 'port', 'type', 'balance', 'expiry'),
            show='headings',
            yscrollcommand=scroll_y.set,
            xscrollcommand=scroll_x.set,
            selectmode='none',
            height=6
        )
        self.blacklist_lic_tree.pack(fill='both', expand=True)
        scroll_y.config(command=self.blacklist_lic_tree.yview)
        scroll_x.config(command=self.blacklist_lic_tree.xview)
        
        self.blacklist_lic_tree.heading('selected', text="[Select]")
        self.blacklist_lic_tree.heading('lic_no', text="License No.")
        self.blacklist_lic_tree.heading('port', text="Port")
        self.blacklist_lic_tree.heading('type', text="Type")
        self.blacklist_lic_tree.heading('balance', text="Balance (INR)")
        self.blacklist_lic_tree.heading('expiry', text="Expiry Date")
        
        self.blacklist_lic_tree.column('selected', width=70, anchor='center')
        self.blacklist_lic_tree.column('lic_no', width=130, anchor='center')
        self.blacklist_lic_tree.column('port', width=80, anchor='center')
        self.blacklist_lic_tree.column('type', width=90, anchor='center')
        self.blacklist_lic_tree.column('balance', width=120, anchor='e')
        self.blacklist_lic_tree.column('expiry', width=100, anchor='center')
        
        self.blacklist_lic_tree.bind("<ButtonRelease-1>", self.on_blacklist_tree_click)
        self.blacklist_lic_tree.tag_configure('checked', background="#E8F5E9")
        
        # Card 2: Proposed Re-allocation Preview Table
        ttk.Label(right_frame, text="Proposed Re-allocation Details Preview", style='Section.TLabel').pack(anchor='w', pady=(10, 5))
        
        self.blacklist_proposed_frame = ttk.Frame(right_frame)
        self.blacklist_proposed_frame.pack(fill='both', expand=True)
        
        prop_scroll_y = ttk.Scrollbar(self.blacklist_proposed_frame, orient='vertical')
        prop_scroll_y.pack(side='right', fill='y')
        
        self.blacklist_proposed_tree = ttk.Treeview(
            self.blacklist_proposed_frame,
            columns=('inv_sr', 'item_sr', 'desc', 'duty', 'new_lic', 'port_match'),
            show='headings',
            yscrollcommand=prop_scroll_y.set,
            height=7
        )
        self.blacklist_proposed_tree.pack(fill='both', expand=True)
        prop_scroll_y.config(command=self.blacklist_proposed_tree.yview)
        
        self.blacklist_proposed_tree.heading('inv_sr', text="Inv Sr.")
        self.blacklist_proposed_tree.heading('item_sr', text="Item Sr.")
        self.blacklist_proposed_tree.heading('desc', text="Description")
        self.blacklist_proposed_tree.heading('duty', text="Duty to Reallocate (INR)")
        self.blacklist_proposed_tree.heading('new_lic', text="Reallocated License")
        self.blacklist_proposed_tree.heading('port_match', text="Port Match?")
        
        self.blacklist_proposed_tree.column('inv_sr', width=60, anchor='center')
        self.blacklist_proposed_tree.column('item_sr', width=60, anchor='center')
        self.blacklist_proposed_tree.column('desc', width=180, anchor='w')
        self.blacklist_proposed_tree.column('duty', width=140, anchor='e')
        self.blacklist_proposed_tree.column('new_lic', width=130, anchor='center')
        self.blacklist_proposed_tree.column('port_match', width=95, anchor='center')
        
        self.blacklist_proposed_tree.tag_configure('error', foreground=COLOR_ERROR)
        self.blacklist_proposed_tree.tag_configure('success', foreground=COLOR_ACCENT)

    def build_rectify_tab(self, parent):
        """Build the layout and widgets for Tab 4: Rectify License."""
        main_pane = ttk.PanedWindow(parent, orient='horizontal')
        main_pane.pack(fill='both', expand=True, padx=10, pady=10)
        
        # --- LEFT SIDE: Blacklisted licenses table ---
        left_frame = ttk.Frame(main_pane, style='Card.TFrame', padding=15)
        main_pane.add(left_frame, weight=1)
        
        ttk.Label(left_frame, text="Blacklisted Licenses", style='Section.TLabel').pack(anchor='w', pady=(0, 15))
        
        # Treeview list
        table_frame = ttk.Frame(left_frame)
        table_frame.pack(fill='both', expand=True, pady=(0, 10))
        
        self.blacklist_tree = ttk.Treeview(table_frame, columns=('lic_no', 'date', 'reason'), show='headings', selectmode='browse')
        self.blacklist_tree.heading('lic_no', text='License No.')
        self.blacklist_tree.heading('date', text='Date Blacklisted')
        self.blacklist_tree.heading('reason', text='Reason')
        
        self.blacklist_tree.column('lic_no', width=130, anchor='center')
        self.blacklist_tree.column('date', width=130, anchor='center')
        self.blacklist_tree.column('reason', width=180, anchor='w')
        
        scrollbar = ttk.Scrollbar(table_frame, orient='vertical', command=self.blacklist_tree.yview)
        self.blacklist_tree.configure(yscrollcommand=scrollbar.set)
        
        self.blacklist_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        self.blacklist_tree.bind('<<TreeviewSelect>>', self.on_blacklist_row_selected)
        
        # Refresh Button
        ttk.Button(left_frame, text="REFRESH BLACKLIST FROM CLOUD", command=self.refresh_blacklist).pack(fill='x')
        
        # --- RIGHT SIDE: Rectification Editor ---
        right_frame = ttk.Frame(main_pane, style='Card.TFrame', padding=15)
        main_pane.add(right_frame, weight=1)
        
        ttk.Label(right_frame, text="Rectification Editor", style='Section.TLabel').pack(anchor='w', pady=(0, 15))
        ttk.Label(right_frame, text="Select a blacklisted license on the left to edit and restore to active pool.", font=(FONT_FAMILY, 9, 'italic'), foreground="#666666").pack(anchor='w', pady=(0, 15))
        
        self.rectify_editor_frame = ttk.Frame(right_frame, style='Card.TFrame')
        self.rectify_editor_frame.pack(fill='x', pady=10)
        
        grid_frame = ttk.Frame(self.rectify_editor_frame, style='Card.TFrame')
        grid_frame.pack(fill='x')
        
        self.rectify_fields = {}
        fields_config = [
            ("License No.:", "lic_no", True),
            ("License Date:", "date", False),
            ("Port of Reg.:", "port", False),
            ("Lic Type:", "type", False),
            ("Value (INR):", "val", False),
            ("Expiry Date:", "expiry", False)
        ]
        
        for i, (label_text, key, is_ro) in enumerate(fields_config):
            row = i // 2
            col = (i % 2) * 2
            
            ttk.Label(grid_frame, text=label_text, style='Summary.TLabel').grid(row=row, column=col, sticky='w', padx=5, pady=8)
            val_var = tk.StringVar()
            self.rectify_fields[key] = val_var
            
            entry = ttk.Entry(grid_frame, textvariable=val_var, width=18)
            if is_ro:
                entry.config(state='readonly')
            entry.grid(row=row, column=col+1, sticky='w', padx=5, pady=8)
            
        self.rectify_restore_btn = ttk.Button(right_frame, text="RECTIFY & RESTORE TO ACTIVE POOL", command=self.run_rectify_restore, state='disabled', style='Action.TButton')
        self.rectify_restore_btn.pack(fill='x', pady=20)

    def browse_blacklist_jd(self):
        filename = filedialog.askopenfilename(title="Select Processed JobData Excel", filetypes=[("Excel Files", "*.xlsx")])
        if filename:
            self.blacklist_jd_path.set(filename)
            self.parse_blacklist_jd()

    def parse_blacklist_jd(self):
        jd_path = self.blacklist_jd_path.get().strip()
        if not jd_path:
            messagebox.showerror("Error", "Please select a Processed JobData Excel file first.")
            return
        if not os.path.exists(jd_path):
            messagebox.showerror("Error", "Selected file does not exist.")
            return
            
        self.root.config(cursor="watch")
        self.blacklist_parse_btn.config(state='disabled')
        self.log("Parsing debited licenses from processed JobData file...")
        
        threading.Thread(target=self._parse_blacklist_jd_worker, args=(jd_path,), daemon=True).start()

    def _parse_blacklist_jd_worker(self, jd_path):
        try:
            wb = openpyxl.load_workbook(jd_path, read_only=True)
            if 'LICENSE' not in wb.sheetnames:
                wb.close()
                raise ValueError("This Excel file is not a processed JobData file (missing the 'LICENSE' sheet).")
                
            ws_lic = wb['LICENSE']
            lics = set()
            for r in list(ws_lic.iter_rows(values_only=True))[1:]:
                if len(r) > 3 and r[3]:
                    lics.add(str(r[3]).strip())
            wb.close()
            
            if not lics:
                raise ValueError("No licenses found in the 'LICENSE' sheet of this JobData file.")
                
            # Extract job no from filename (preferred) or GENERAL sheet
            job_no = extract_job_no_from_filename(jd_path)
            
            import_port = ""
            wb_gen = openpyxl.load_workbook(jd_path, read_only=True)
            if 'GENERAL' in wb_gen.sheetnames:
                ws_gen = wb_gen['GENERAL']
                rows = list(ws_gen.iter_rows(values_only=True))
                if len(rows) > 1:
                    headers = [str(h).strip().upper() for h in rows[0]]
                    if 'CUSTOMSHOUSECODE' in headers:
                        import_port = str(rows[1][headers.index('CUSTOMSHOUSECODE')]).strip()
            wb_gen.close()
            
            sorted_lics = sorted(list(lics))
            self.gui_queue.put(lambda: self._parse_blacklist_jd_success(job_no, import_port, sorted_lics))
            
        except Exception as e:
            err_msg = str(e)
            self.gui_queue.put(lambda: self._parse_blacklist_jd_failure(err_msg))

    def _parse_blacklist_jd_success(self, job_no, import_port, sorted_lics):
        self.root.config(cursor="")
        self.blacklist_parse_btn.config(state='normal')
        
        self.blacklist_job_info = {
            'job_no': job_no,
            'import_port': import_port
        }
        
        self.blacklist_parsed_debits = []
        self.blacklist_selected_to_blacklist.clear()
        
        for lic_no in sorted_lics:
            matched = next((l for l in self.active_licenses if l['lic_no'] == lic_no), None)
            if not matched:
                matched = next((l for l in self.used_licenses if l['lic_no'] == lic_no), None)
            l_port = matched['port'] if matched else "N/A"
            l_type = matched['type'] if matched else "N/A"
            
            self.blacklist_parsed_debits.append({
                'lic_no': lic_no,
                'type': l_type,
                'port': l_port
            })
            
            pass
            
        # Reset search field and populate tree
        self.blacklist_search_var.set("")
        self.populate_blacklist_select_tree()
        self.on_blacklist_lic_changed()
        
        self.log(f"Parsed {len(sorted_lics)} licenses from Job '{job_no}' successfully.")

    def _parse_blacklist_jd_failure(self, error_msg):
        self.root.config(cursor="")
        self.blacklist_parse_btn.config(state='normal')
        self.blacklist_execute_btn.config(state='disabled')
        
        self.blacklist_parsed_debits = []
        self.blacklist_selected_to_blacklist.clear()
        self.blacklist_search_var.set("")
        
        for item_id in self.blacklist_select_tree.get_children():
            self.blacklist_select_tree.delete(item_id)
            
        messagebox.showerror("Error", f"Failed to parse JobData file:\n\n{error_msg}", parent=self.root)

    def on_blacklist_select_tree_click(self, event):
        item_id = self.blacklist_select_tree.identify_row(event.y)
        if not item_id:
            return
            
        values = list(self.blacklist_select_tree.item(item_id, 'values'))
        lic_no = values[1]
        if values[0] == "☑":
            values[0] = "☐"
            self.blacklist_selected_to_blacklist.discard(lic_no)
        else:
            values[0] = "☑"
            self.blacklist_selected_to_blacklist.add(lic_no)
            
        current_tags = list(self.blacklist_select_tree.item(item_id, 'tags'))
        if values[0] == "☑":
            if 'checked' not in current_tags:
                current_tags.append('checked')
        else:
            if 'checked' in current_tags:
                current_tags.remove('checked')
                
        self.blacklist_select_tree.item(item_id, values=values, tags=tuple(current_tags))
        self.on_blacklist_lic_changed()

    def populate_blacklist_select_tree(self):
        """Populate the select licenses to blacklist treeview with filtering."""
        for item_id in self.blacklist_select_tree.get_children():
            self.blacklist_select_tree.delete(item_id)
            
        query = self.blacklist_search_var.get().strip().lower()
        
        for item in self.blacklist_parsed_debits:
            lic_no = item['lic_no']
            l_type = item['type']
            l_port = item['port']
            
            # Simple substring matching for search entry
            if query:
                if query not in lic_no.lower() and query not in l_type.lower() and query not in l_port.lower():
                    continue
                    
            is_checked = lic_no in self.blacklist_selected_to_blacklist
            select_str = "☑" if is_checked else "☐"
            tags = ('checked',) if is_checked else ()
            
            self.blacklist_select_tree.insert(
                '', 
                'end', 
                values=(select_str, lic_no, l_type, l_port),
                tags=tags
            )

    def on_blacklist_dropdown_checker_changed(self, event):
        """Handle Check All / Uncheck All choices from the dropdown checker on the visible rows."""
        sel = self.blacklist_dropdown_checker.get()
        if not sel:
            return
            
        # Collect currently visible lic_no strings from the treeview
        visible_lics = []
        for item_id in self.blacklist_select_tree.get_children():
            values = self.blacklist_select_tree.item(item_id, 'values')
            visible_lics.append(values[1])
            
        if sel == "Check All":
            for lic_no in visible_lics:
                self.blacklist_selected_to_blacklist.add(lic_no)
        elif sel == "Uncheck All":
            for lic_no in visible_lics:
                self.blacklist_selected_to_blacklist.discard(lic_no)
                
        # Reset combo value to select placeholder
        self.blacklist_dropdown_checker.set("[Select All/None]")
        
        # Redraw selection treeview and trigger balance preview update
        self.populate_blacklist_select_tree()
        self.on_blacklist_lic_changed()

    def on_blacklist_lic_changed(self):
        jd_path = self.blacklist_jd_path.get().strip()
        if not (self.blacklist_selected_to_blacklist and jd_path):
            # Clear preview treeviews and label
            for item_id in self.blacklist_lic_tree.get_children():
                self.blacklist_lic_tree.delete(item_id)
            for item_id in self.blacklist_proposed_tree.get_children():
                self.blacklist_proposed_tree.delete(item_id)
            self.blacklist_summary_lbl.config(text="Please select one or more licenses to blacklist.")
            self.blacklist_execute_btn.config(state='disabled')
            return
            
        self.root.config(cursor="watch")
        lics_list = list(self.blacklist_selected_to_blacklist)
        self.log(f"Loading affected items for licenses {', '.join(lics_list)}...")
        threading.Thread(target=self._load_affected_items_worker, args=(jd_path, lics_list), daemon=True).start()

    def _load_affected_items_worker(self, jd_path, lics_to_blacklist):
        try:
            wb_jd = openpyxl.load_workbook(jd_path, data_only=True)
            ws_lic = wb_jd['LICENSE']
            ws_items = wb_jd['ITEMS']
            
            # Read items descriptions from ITEMS sheet
            items_rows = list(ws_items.iter_rows(values_only=True))
            items_header = [str(h).strip() for h in items_rows[0]]
            item_inv_sr_idx = items_header.index('InvSrNo')
            item_sr_idx = items_header.index('ItemSrNo')
            desc_idx = items_header.index('Product_Description')
            cth_idx = items_header.index('CTH')
            exim_notn_idx = -1
            for name in ['Exim_Notn', 'Exim_Notification']:
                if name in items_header:
                    exim_notn_idx = items_header.index(name)
                    break
            
            item_desc_map = {}
            item_exim_notn_map = {}
            for r in items_rows[1:]:
                if r[item_inv_sr_idx] is not None and r[item_sr_idx] is not None:
                    k = (str(r[item_inv_sr_idx]).strip(), str(r[item_sr_idx]).strip())
                    item_desc_map[k] = {
                        'desc': str(r[desc_idx]).strip() if r[desc_idx] is not None else "",
                        'cth': str(r[cth_idx]).strip() if r[cth_idx] is not None else "",
                    }
                    if exim_notn_idx != -1:
                        item_exim_notn_map[k] = str(r[exim_notn_idx]).strip().upper() if r[exim_notn_idx] is not None else ""
            
            # Find Invoice No mapping
            ws_inv = wb_jd['INVOICES']
            inv_rows = list(ws_inv.iter_rows(values_only=True))
            inv_header = [str(h).strip() for h in inv_rows[0]]
            inv_sr_idx = inv_header.index('InvSrNo')
            inv_no_idx = inv_header.index('Invoice_No')
            inv_map = {}
            for r in inv_rows[1:]:
                if r[inv_sr_idx] is not None and r[inv_no_idx] is not None:
                    inv_map[str(r[inv_sr_idx]).strip()] = str(r[inv_no_idx]).strip()
            
            # Collect debits of the licenses to blacklist
            lic_rows_raw = list(ws_lic.iter_rows(values_only=True))
            
            # Calculate total duty and quantity of each item across all its allocations in the LICENSE sheet
            item_total_duty_map = {}
            item_total_qty_map = {}
            for row in lic_rows_raw[1:]:
                if len(row) > 11 and row[0] is not None and row[1] is not None:
                    inv_sr_val = str(row[0]).strip()
                    item_sr_val = str(row[1]).strip()
                    key_val = (inv_sr_val, item_sr_val)
                    item_total_duty_map[key_val] = item_total_duty_map.get(key_val, 0.0) + safe_float(row[10])
                    item_total_qty_map[key_val] = item_total_qty_map.get(key_val, 0.0) + safe_float(row[11])

            affected_items = []
            lics_set = set(lics_to_blacklist)
            
            for row in lic_rows_raw[1:]:
                if len(row) > 3 and row[3]:
                    lic_num = str(row[3]).strip()
                    if lic_num in lics_set:
                        inv_sr = str(row[0]).strip()
                        item_sr = str(row[1]).strip()
                        k = (inv_sr, item_sr)
                        desc_info = item_desc_map.get(k, {'desc': '', 'cth': ''})
                        
                        affected_items.append({
                            'Inv_SrNo': inv_sr,
                            'Item_SrNo': item_sr,
                            'Invoice_No': inv_map.get(inv_sr, f"Inv-{inv_sr}"),
                            'License_No': lic_num,
                            'License_Date': row[4],
                            'License_RegNo': row[5],
                            'License_RegDate': row[6],
                            'Reg_Port': row[7],
                            'CIF_Value': safe_float(row[9]),
                            'DebitDeutyValue': safe_float(row[10]),
                            'DebitQuantity': safe_float(row[11]),
                            'DebitQuantityUnitCode': row[12] if len(row) > 12 else "",
                            'Product_Desc': desc_info['desc'],
                            'CTH': desc_info['cth'],
                            'Total_Item_Duty': item_total_duty_map.get(k, 0.0),
                            'Total_Item_Qty': item_total_qty_map.get(k, 0.0)
                        })
            
            wb_jd.close()
            
            # Determine the type of the blacklisted license
            lic_type = "RODTEP" # Fallback
            for alloc in affected_items:
                k = (alloc['Inv_SrNo'], alloc['Item_SrNo'])
                notn = item_exim_notn_map.get(k, "")
                allowed = "RODTEP"
                if 'ROSCTL' in notn:
                    allowed = "ROSCTL"
                    lic_type = "ROSCTL"
                elif 'DFIA' in notn:
                    allowed = "DFIA"
                    lic_type = "DFIA"
                elif 'EPCG' in notn:
                    allowed = "EPCG"
                    lic_type = "EPCG"
                elif 'ADVANCE' in notn or 'ADV' in notn or 'AA' in notn:
                    allowed = "ADVANCE"
                    lic_type = "ADVANCE"
                alloc['Allowed_Type'] = allowed
            
            self.gui_queue.put(lambda: self._load_affected_items_success(affected_items, lic_type))
        except Exception as e:
            err_msg = str(e)
            self.gui_queue.put(lambda: self._load_affected_items_failure(err_msg))

    def _load_affected_items_success(self, affected_items, lic_type):
        self.root.config(cursor="")
        self.blacklist_affected_items = affected_items
        self.blacklist_job_info['lic_type'] = lic_type
        
        # Populate all active licenses except the blacklisted ones (allows multiple types)
        active_lics = []
        for l in self.active_licenses:
            if l.get('lic_no') in self.blacklist_selected_to_blacklist:
                continue
            active_lics.append(l)
            
        self.blacklist_active_lics = active_lics
        
        # Run simulation to see which licenses are actually needed to cover reallocation duty
        # Each item duty must be fully covered by a single type, trying types in top-to-bottom order.
        for l in active_lics:
            l['sim_bal'] = l['bal']
            l['sim_used'] = False
            
        for alloc in affected_items:
            duty = alloc['DebitDeutyValue']
            if duty <= 0:
                if active_lics:
                    active_lics[0]['sim_used'] = True
                continue
                
            allocated_type = None
            tried_types = set()
            for lic in active_lics:
                lic_type_name = str(lic['type']).strip().upper()
                if lic_type_name in tried_types:
                    continue
                max_start = lic['sim_bal'] - 1.00
                if max_start <= 0.01:
                    continue
                total_avail = sum(
                    max(0.0, l['sim_bal'] - 1.00)
                    for l in active_lics
                    if str(l['type']).strip().upper() == lic_type_name
                )
                if total_avail >= duty:
                    allocated_type = lic_type_name
                    break
                else:
                    tried_types.add(lic_type_name)
                    
            if not allocated_type:
                continue
                
            duty_remaining = duty
            for lic in active_lics:
                if duty_remaining <= 0.005:
                    break
                if str(lic['type']).strip().upper() != allocated_type:
                    continue
                max_debitable = lic['sim_bal'] - 1.00
                if max_debitable <= 0.01:
                    continue
                if max_debitable >= duty_remaining:
                    lic['sim_bal'] -= duty_remaining
                    lic['sim_used'] = True
                    duty_remaining = 0
                    break
                else:
                    debit_amt = max_debitable
                    lic['sim_bal'] = 1.00
                    lic['sim_used'] = True
                    duty_remaining -= debit_amt
                    
        # Clear active licenses treeview
        for item_id in self.blacklist_lic_tree.get_children():
            self.blacklist_lic_tree.delete(item_id)
            
        self.blacklist_selected_lic_nos.clear()
        
        # Populate Active Licenses in Treeview, checking only the ones marked as sim_used
        for l in active_lics:
            expiry_str = l['expiry'].strftime('%d-%b-%Y') if l['expiry'] else ""
            if l['sim_used']:
                self.blacklist_lic_tree.insert(
                    '', 
                    'end', 
                    values=("☑", l['lic_no'], l['port'], l['type'], f"{l['bal']:,.2f}", expiry_str),
                    tags=('checked',)
                )
                self.blacklist_selected_lic_nos.add(l['lic_no'])
            else:
                self.blacklist_lic_tree.insert(
                    '', 
                    'end', 
                    values=("☐", l['lic_no'], l['port'], l['type'], f"{l['bal']:,.2f}", expiry_str)
                )
            
        self.recalc_blacklist_reallocation_metrics()

    def _load_affected_items_failure(self, error_msg):
        self.root.config(cursor="")
        messagebox.showerror("Error", f"Failed to load affected items details:\n\n{error_msg}", parent=self.root)

    def on_blacklist_tree_click(self, event):
        """Toggle checkbox status on blacklist active licenses treeview row click."""
        item_id = self.blacklist_lic_tree.identify_row(event.y)
        if not item_id:
            return
            
        values = list(self.blacklist_lic_tree.item(item_id, 'values'))
        if values[0] == "☑":
            values[0] = "☐"
            self.blacklist_selected_lic_nos.discard(values[1])
        else:
            values[0] = "☑"
            self.blacklist_selected_lic_nos.add(values[1])
            
        current_tags = list(self.blacklist_lic_tree.item(item_id, 'tags'))
        if values[0] == "☑":
            if 'checked' not in current_tags:
                current_tags.append('checked')
        else:
            if 'checked' in current_tags:
                current_tags.remove('checked')
                
        self.blacklist_lic_tree.item(item_id, values=values, tags=tuple(current_tags))
        self.recalc_blacklist_reallocation_metrics()

    def recalc_blacklist_reallocation_metrics(self):
        """Re-simulate reallocation using the currently selected active licenses pool."""
        selected_lics = []
        for item_id in self.blacklist_lic_tree.get_children():
            values = self.blacklist_lic_tree.item(item_id, 'values')
            if values[0] == "☑":
                lic_no = values[1]
                matched_lic = next((l for l in self.blacklist_active_lics if l['lic_no'] == lic_no), None)
                if matched_lic:
                    selected_lics.append({
                        'lic_no': matched_lic['lic_no'],
                        'port': matched_lic['port'],
                        'type': matched_lic['type'],
                        'val': matched_lic['val'],
                        'bal': matched_lic['bal'],
                        'reg_date': matched_lic['reg_date'],
                        'expiry': matched_lic['expiry'],
                        'sim_bal': matched_lic['bal']
                    })
                    
        # Sort using same priority rules: port matches first, then expiry date
        import_port = self.blacklist_job_info.get('import_port', "")
        # Preserve strict database sheet (top-to-bottom) order without sorting
        sorted_lics = selected_lics
        
        reallocations_preview = []
        all_covered = True
        swapped_items = set()
        
        for alloc in self.blacklist_affected_items:
            inv_sr = alloc['Inv_SrNo']
            item_sr = alloc['Item_SrNo']
            k = (inv_sr, item_sr)
            if k in swapped_items:
                continue
                
            duty = alloc['DebitDeutyValue']
            qty = alloc['DebitQuantity']
            av = alloc['CIF_Value']
            old_lic = alloc.get('License_No', '')
            
            if duty <= 0:
                lic = sorted_lics[0] if sorted_lics else None
                if lic:
                    reallocations_preview.append({
                        'Inv_SrNo': inv_sr,
                        'Item_SrNo': item_sr,
                        'Invoice_No': alloc['Invoice_No'],
                        'Old_License_No': old_lic,
                        'License_No': lic['lic_no'],
                        'License_Type': lic['type'],
                        'DebitDeutyValue': 0.00,
                        'DebitQuantity': qty,
                        'Product_Desc': alloc['Product_Desc'],
                        'CTH': alloc['CTH'],
                        'port_match': "YES" if lic['port'] == import_port else "NO",
                        'status': 'success'
                    })
                else:
                    reallocations_preview.append({
                        'Inv_SrNo': inv_sr,
                        'Item_SrNo': item_sr,
                        'Invoice_No': alloc['Invoice_No'],
                        'Old_License_No': old_lic,
                        'License_No': "NO LICENSE SELECTED",
                        'License_Type': "Unknown",
                        'DebitDeutyValue': 0.00,
                        'DebitQuantity': qty,
                        'Product_Desc': alloc['Product_Desc'],
                        'CTH': alloc['CTH'],
                        'port_match': "NO",
                        'status': 'error'
                    })
                    all_covered = False
                continue
                
            # Priority 1: Same type reallocation
            # Find which license of the SAME type can cover this portion
            allocated_type = None
            tried_types = set()
            allowed_type_str = str(alloc.get('Allowed_Type', '')).strip().upper()
            
            for lic in sorted_lics:
                lic_type_name = str(lic['type']).strip().upper()
                if lic_type_name in tried_types:
                    continue
                
                # Check: Only allow this license type if it matches the item's Allowed_Type
                is_match = True
                if allowed_type_str:
                    is_match = False
                    if 'ROSCTL' in allowed_type_str and 'ROSCTL' in lic_type_name:
                        is_match = True
                    elif ('RODTEP' in allowed_type_str or 'RD' in allowed_type_str) and ('RODTEP' in lic_type_name or 'RD' in lic_type_name):
                        is_match = True
                    elif 'DFIA' in allowed_type_str and 'DFIA' in lic_type_name:
                        is_match = True
                    elif 'EPCG' in allowed_type_str and 'EPCG' in lic_type_name:
                        is_match = True
                    elif ('ADVANCE' in allowed_type_str or 'ADV' in allowed_type_str or 'AA' in allowed_type_str) and ('ADVANCE' in lic_type_name or 'ADV' in lic_type_name or 'AA' in lic_type_name):
                        is_match = True
                        
                if is_match:
                    max_start = lic['sim_bal'] - 1.00
                    if max_start <= 0.01:
                        continue
                        
                    total_avail = sum(
                        max(0.0, l['sim_bal'] - 1.00)
                        for l in sorted_lics
                        if str(l['type']).strip().upper() == lic_type_name
                    )
                    if total_avail >= duty:
                        allocated_type = lic_type_name
                        break
                    else:
                        tried_types.add(lic_type_name)
                        
            # Priority 2: Different type swap for the entire item
            swap_lic = None
            if not allocated_type:
                total_item_duty = alloc.get('Total_Item_Duty', duty)
                for lic in sorted_lics:
                    lic_type_name = str(lic['type']).strip().upper()
                    
                    # Must be a DIFFERENT type
                    is_same = False
                    if allowed_type_str:
                        if 'ROSCTL' in allowed_type_str and 'ROSCTL' in lic_type_name:
                            is_same = True
                        elif ('RODTEP' in allowed_type_str or 'RD' in allowed_type_str) and ('RODTEP' in lic_type_name or 'RD' in lic_type_name):
                            is_same = True
                        elif 'DFIA' in allowed_type_str and 'DFIA' in lic_type_name:
                            is_same = True
                        elif 'EPCG' in allowed_type_str and 'EPCG' in lic_type_name:
                            is_same = True
                        elif ('ADVANCE' in allowed_type_str or 'ADV' in allowed_type_str or 'AA' in allowed_type_str) and ('ADVANCE' in lic_type_name or 'ADV' in lic_type_name or 'AA' in lic_type_name):
                            is_same = True
                            
                    if is_same:
                        continue
                        
                    # Must have enough balance for the entire item's duty
                    if lic['sim_bal'] - 1.00 >= total_item_duty:
                        swap_lic = lic
                        break
                        
            if swap_lic:
                swap_lic['sim_bal'] -= total_item_duty
                swapped_items.add(k)
                reallocations_preview.append({
                    'Inv_SrNo': inv_sr,
                    'Item_SrNo': item_sr,
                    'Invoice_No': alloc['Invoice_No'],
                    'Old_License_No': old_lic,
                    'License_No': swap_lic['lic_no'],
                    'License_Type': swap_lic['type'],
                    'DebitDeutyValue': round(total_item_duty, 2),
                    'DebitQuantity': alloc.get('Total_Item_Qty', qty),
                    'Product_Desc': alloc['Product_Desc'],
                    'CTH': alloc['CTH'],
                    'port_match': "YES" if swap_lic['port'] == import_port else "NO",
                    'status': 'success',
                    'is_swap': True
                })
                continue
                
            if not allocated_type and not swap_lic:
                reallocations_preview.append({
                    'Inv_SrNo': inv_sr,
                    'Item_SrNo': item_sr,
                    'Invoice_No': alloc['Invoice_No'],
                    'Old_License_No': old_lic,
                    'License_No': "UNALLOCATED (LACKS BAL)",
                    'License_Type': "Unknown",
                    'DebitDeutyValue': round(duty, 2),
                    'DebitQuantity': qty,
                    'Product_Desc': alloc['Product_Desc'],
                    'CTH': alloc['CTH'],
                    'port_match': "NO",
                    'status': 'error'
                })
                all_covered = False
                continue
                
            # Allocate from selected licenses of this type (Priority 1)
            duty_remaining = duty
            for lic in sorted_lics:
                if duty_remaining <= 0.005:
                    break
                if str(lic['type']).strip().upper() != allocated_type:
                    continue
                max_debitable = lic['sim_bal'] - 1.00
                if max_debitable <= 0.01:
                    continue
                    
                if max_debitable >= duty_remaining:
                    lic['sim_bal'] -= duty_remaining
                    reallocations_preview.append({
                        'Inv_SrNo': inv_sr,
                        'Item_SrNo': item_sr,
                        'Invoice_No': alloc['Invoice_No'],
                        'Old_License_No': old_lic,
                        'License_No': lic['lic_no'],
                        'License_Type': lic['type'],
                        'DebitDeutyValue': round(duty_remaining, 2),
                        'DebitQuantity': qty,
                        'Product_Desc': alloc['Product_Desc'],
                        'CTH': alloc['CTH'],
                        'port_match': "YES" if lic['port'] == import_port else "NO",
                        'status': 'success'
                    })
                    duty_remaining = 0
                    break
                else:
                    debit_amt = max_debitable
                    lic['sim_bal'] = 1.00
                    ratio = debit_amt / duty
                    reallocations_preview.append({
                        'Inv_SrNo': inv_sr,
                        'Item_SrNo': item_sr,
                        'Invoice_No': alloc['Invoice_No'],
                        'Old_License_No': old_lic,
                        'License_No': lic['lic_no'],
                        'License_Type': lic['type'],
                        'DebitDeutyValue': round(debit_amt, 2),
                        'DebitQuantity': round(qty * ratio, 3),
                        'Product_Desc': alloc['Product_Desc'],
                        'CTH': alloc['CTH'],
                        'port_match': "YES" if lic['port'] == import_port else "NO",
                        'status': 'success'
                    })
                    duty_remaining -= debit_amt
                    
        # Update proposed treeview preview
        for item_id in self.blacklist_proposed_tree.get_children():
            self.blacklist_proposed_tree.delete(item_id)
            
        for p in reallocations_preview:
            tag = p['status']
            self.blacklist_proposed_tree.insert(
                '', 
                'end', 
                values=(
                    p['Invoice_No'],
                    p['Item_SrNo'],
                    p['Product_Desc'],
                    f"{p['DebitDeutyValue']:,.2f}",
                    p['License_No'],
                    p['port_match']
                ),
                tags=(tag,)
            )
            
        # Update details summary text
        job_no = self.blacklist_job_info.get('job_no', 'Unknown')
        lics_to_bl = ", ".join(self.blacklist_selected_to_blacklist)
        lic_type = self.blacklist_job_info.get('lic_type', 'Unknown')
        total_duty = sum(a['DebitDeutyValue'] for a in self.blacklist_affected_items)
        covered_duty = sum(p['DebitDeutyValue'] for p in reallocations_preview if p['status'] == 'success')
        
        status_text = (
            f"Job Number: {job_no}\n"
            f"Import Port: {import_port}\n"
            f"Licenses to Blacklist: {lics_to_bl} ({lic_type})\n"
            f"Total Reallocation Duty: {total_duty:,.2f} INR\n"
            f"Duty Covered: {covered_duty:,.2f} INR\n"
            f"Status: "
        )
        if total_duty > 0:
            if covered_duty > 0:
                if all_covered:
                    status_text += "READY FOR REALLOCATION"
                else:
                    status_text += "PARTIAL REALLOCATION READY (Some items will be skipped due to low balance)"
            else:
                status_text += "READY FOR REALLOCATION (All items will be unallocated / cleared to pending)"
            self.blacklist_execute_btn.config(state='normal')
        else:
            status_text += "No items to reallocate."
            self.blacklist_execute_btn.config(state='disabled')
            
        self.blacklist_summary_lbl.config(text=status_text)
        self.blacklist_reallocations_preview = reallocations_preview

    def run_blacklist_reallocation(self):
        jd_path = self.blacklist_jd_path.get().strip()
        lics_to_blacklist = list(self.blacklist_selected_to_blacklist)
        reason = self.blacklist_reason_entry.get().strip()
        
        if not jd_path or not lics_to_blacklist:
            messagebox.showerror("Error", "Configure JobData and select one or more licenses to blacklist.", parent=self.root)
            return
            
        if not reason:
            if not messagebox.askyesno("Confirm", "You have not provided a reason for blacklisting. Proceed anyway?", parent=self.root):
                return
                
        self.root.config(cursor="watch")
        self.blacklist_execute_btn.config(state='disabled')
        self.blacklist_parse_btn.config(state='disabled')
        self.log(f"Initiating Blacklist & Re-allocation of licenses {', '.join(lics_to_blacklist)}...")
        
        threading.Thread(target=self._blacklist_reallocation_worker, args=(jd_path, lics_to_blacklist, reason), daemon=True).start()

    def _blacklist_reallocation_worker(self, jd_path, lics_to_blacklist, reason):
        try:
            job_no = self.blacklist_job_info.get('job_no', 'Unknown')
            lic_type = self.blacklist_job_info.get('lic_type', 'Unknown')
            
            lic_details_map = {l['lic_no']: l for l in self.blacklist_active_lics}
            lics_set = set(lics_to_blacklist)
            
            swapped_items_map = {}
            reallocated_rows = []
            reallocations_api = []
            skipped_reallocations = []
            failed_keys = set()
            
            for p in self.blacklist_reallocations_preview:
                old_lic = p.get('Old_License_No', '')
                match_affected = next(
                    (a for a in self.blacklist_affected_items 
                     if a['Inv_SrNo'] == p['Inv_SrNo'] 
                     and a['Item_SrNo'] == p['Item_SrNo'] 
                     and a['License_No'] == old_lic), 
                    {}
                )
                
                if p['status'] == 'success':
                    lic_no = p['License_No']
                    lic_obj = lic_details_map.get(lic_no, {})
                    
                    reg_date_str = lic_obj.get('reg_date').strftime('%d-%b-%Y') if lic_obj.get('reg_date') else ""
                    port_str = lic_obj.get('port', "")
                    
                    is_swap = p.get('is_swap', False)
                    if is_swap:
                        swapped_items_map[(p['Inv_SrNo'], p['Item_SrNo'])] = lic_no
                        reallocated_rows.append({
                            'Inv_SrNo': p['Inv_SrNo'],
                            'Item_SrNo': p['Item_SrNo'],
                            'License_No': lic_no,
                            'License_Date': reg_date_str,
                            'License_RegNo': lic_no,
                            'License_RegDate': reg_date_str,
                            'Reg_Port': port_str,
                            'CIF_Value': match_affected.get('CIF_Value', 0.0),
                            'DebitDeutyValue': p['DebitDeutyValue'],
                            'DebitQuantity': p['DebitQuantity'],
                            'DebitQuantityUnitCode': match_affected.get('DebitQuantityUnitCode', "")
                        })
                        reallocations_api.append({
                            'old_lic': old_lic,
                            'new_lic': lic_no,
                            'val': match_affected.get('DebitDeutyValue', 0.0),
                            'inv_no': p['Invoice_No'],
                            'inv_sr_no': p['Inv_SrNo'],
                            'item_sr_no': p['Item_SrNo'],
                            'desc': p['Product_Desc'],
                            'job_no': job_no
                        })
                    else:
                        reallocated_rows.append({
                            'Inv_SrNo': p['Inv_SrNo'],
                            'Item_SrNo': p['Item_SrNo'],
                            'License_No': lic_no,
                            'License_Date': reg_date_str,
                            'License_RegNo': lic_no,
                            'License_RegDate': reg_date_str,
                            'Reg_Port': port_str,
                            'CIF_Value': match_affected.get('CIF_Value', 0.0),
                            'DebitDeutyValue': p['DebitDeutyValue'],
                            'DebitQuantity': p['DebitQuantity'],
                            'DebitQuantityUnitCode': match_affected.get('DebitQuantityUnitCode', "")
                        })
                        reallocations_api.append({
                            'old_lic': old_lic,
                            'new_lic': lic_no,
                            'val': p['DebitDeutyValue'],
                            'inv_no': p['Invoice_No'],
                            'inv_sr_no': p['Inv_SrNo'],
                            'item_sr_no': p['Item_SrNo'],
                            'desc': p['Product_Desc'],
                            'job_no': job_no
                        })
                else:
                    skipped_reallocations.append(p)
                    failed_keys.add((p['Inv_SrNo'], p['Item_SrNo']))
                    # For skipped items, restore balance on old license, but do not debit any new license (new_lic: "")
                    reallocations_api.append({
                        'old_lic': old_lic,
                        'new_lic': "",
                        'val': p['DebitDeutyValue'],
                        'inv_no': p['Invoice_No'],
                        'inv_sr_no': p['Inv_SrNo'],
                        'item_sr_no': p['Item_SrNo'],
                        'desc': p['Product_Desc'],
                        'job_no': job_no
                    })
                
            # 1. Modify the local processed JobData Excel file
            self.safe_log("Modifying local Excel file...")
            wb_write = openpyxl.load_workbook(jd_path, data_only=False)
            ws_write_lic = wb_write['LICENSE']
            ws_write_items = wb_write['ITEMS']
            
            # Read and keep other license rows, excluding those that belong to failed reallocations!
            other_allocations = []
            lic_rows_raw = list(ws_write_lic.iter_rows(values_only=True))
            
            # Create a lookup for preview details (Invoice_No and Product_Desc) to use for secondary license API payload
            preview_map = {(p['Inv_SrNo'], p['Item_SrNo']): p for p in self.blacklist_reallocations_preview}
            
            for row in lic_rows_raw[1:]:
                if len(row) > 3 and row[3]:
                    row_inv_sr = str(row[0]).strip()
                    row_item_sr = str(row[1]).strip()
                    row_lic_no = str(row[3]).strip()
                    k = (row_inv_sr, row_item_sr)
                    
                    if row_lic_no in lics_set:
                        # This is a blacklisted license row. Ignore (handled above)
                        continue
                        
                    if k in failed_keys:
                        # This item failed reallocation of its blacklisted license!
                        # As per user directive, we MUST remove ALL other licenses from it too.
                        # We do NOT write this row back to the Excel.
                        # And we MUST restore its balance in Google Sheets!
                        p_obj = preview_map.get(k, {})
                        reallocations_api.append({
                            'old_lic': row_lic_no,
                            'new_lic': "",
                            'val': safe_float(row[10]),
                            'inv_no': p_obj.get('Invoice_No', f"Inv-{row_inv_sr}"),
                            'inv_sr_no': row_inv_sr,
                            'item_sr_no': row_item_sr,
                            'desc': p_obj.get('Product_Desc', ''),
                            'job_no': job_no
                        })
                        self.safe_log(f"Removing secondary active license {row_lic_no} from failed item (InvSrNo: {row_inv_sr}, ItemSrNo: {row_item_sr}) to prevent partial allocation.")
                    elif k in swapped_items_map:
                        # This item was swapped to a different type license!
                        # We remove this active license row from the Excel (handled under the single new swap row)
                        # and update Google Sheets by replacing this active license with the new swapped license.
                        p_obj = preview_map.get(k, {})
                        reallocations_api.append({
                            'old_lic': row_lic_no,
                            'new_lic': swapped_items_map[k],
                            'val': safe_float(row[10]),
                            'inv_no': p_obj.get('Invoice_No', f"Inv-{row_inv_sr}"),
                            'inv_sr_no': row_inv_sr,
                            'item_sr_no': row_item_sr,
                            'desc': p_obj.get('Product_Desc', ''),
                            'job_no': job_no
                        })
                        self.safe_log(f"Swapping secondary active license {row_lic_no} to new license {swapped_items_map[k]} for swapped item (InvSrNo: {row_inv_sr}, ItemSrNo: {row_item_sr}).")
                    else:
                        other_allocations.append(row)
                        
            # Clear and rewrite LICENSE sheet
            max_r = ws_write_lic.max_row
            if max_r > 1:
                ws_write_lic.delete_rows(2, max_r)
                
            write_idx = 2
            # Write back other unchanged rows
            for row in other_allocations:
                for c_idx, val in enumerate(row, start=1):
                    ws_write_lic.cell(row=write_idx, column=c_idx, value=val)
                write_idx += 1
                
            # Write back successfully reallocated rows
            for row in reallocated_rows:
                ws_write_lic.cell(row=write_idx, column=1, value=row['Inv_SrNo'])
                ws_write_lic.cell(row=write_idx, column=2, value=row['Item_SrNo'])
                ws_write_lic.cell(row=write_idx, column=3, value=None)
                ws_write_lic.cell(row=write_idx, column=4, value=row['License_No'])
                ws_write_lic.cell(row=write_idx, column=5, value=row['License_Date'])
                ws_write_lic.cell(row=write_idx, column=6, value=row['License_RegNo'])
                ws_write_lic.cell(row=write_idx, column=7, value=row['License_RegDate'])
                ws_write_lic.cell(row=write_idx, column=8, value=row['Reg_Port'])
                ws_write_lic.cell(row=write_idx, column=9, value=1)
                ws_write_lic.cell(row=write_idx, column=10, value=row['CIF_Value'])
                ws_write_lic.cell(row=write_idx, column=11, value=row['DebitDeutyValue'])
                ws_write_lic.cell(row=write_idx, column=12, value=row['DebitQuantity'])
                ws_write_lic.cell(row=write_idx, column=13, value=row['DebitQuantityUnitCode'])
                write_idx += 1
                
            # Update ITEMS sheet notification columns based on new allocations
            item_last_lic = { (r['Inv_SrNo'], r['Item_SrNo']): r['License_No'] for r in reallocated_rows }
            item_last_lic_type = {}
            for p in self.blacklist_reallocations_preview:
                if p['status'] == 'success':
                    item_last_lic_type[(p['Inv_SrNo'], p['Item_SrNo'])] = p.get('License_Type', 'RODTEP')
            
            items_header_write = [str(ws_write_items.cell(row=1, column=c).value).strip() for c in range(1, ws_write_items.max_column + 1)]
            w_inv_sr_idx = items_header_write.index('InvSrNo') + 1
            w_item_sr_idx = items_header_write.index('ItemSrNo') + 1
            
            w_exim_code_idx = items_header_write.index('Exim_Code') + 1 if 'Exim_Code' in items_header_write else -1
            
            w_exim_notn_idx = -1
            for name in ['Exim_Notn', 'Exim_Notification']:
                if name in items_header_write:
                    w_exim_notn_idx = items_header_write.index(name) + 1
                    break
                    
            w_exim_notn_sr_idx = -1
            for name in ['Exim_NotnSrNo', 'Exim_Notification_SrNo']:
                if name in items_header_write:
                    w_exim_notn_sr_idx = items_header_write.index(name) + 1
                    break
                    
            w_duty_exemption_idx = items_header_write.index('Duty_Exemption_Amount') + 1 if 'Duty_Exemption_Amount' in items_header_write else -1
            
            affected_keys = { (a['Inv_SrNo'], a['Item_SrNo']) for a in self.blacklist_affected_items }
            
            for row_idx in range(2, ws_write_items.max_row + 1):
                inv_val = str(ws_write_items.cell(row=row_idx, column=w_inv_sr_idx).value or "").strip()
                itm_val = str(ws_write_items.cell(row=row_idx, column=w_item_sr_idx).value or "").strip()
                k = (inv_val, itm_val)
                
                if k in item_last_lic:
                    new_lic_type = str(item_last_lic_type.get(k, 'RODTEP')).strip().upper()
                    
                    exim_code_val = 'RD'
                    exim_notn_val = 'RODTEP'
                    exim_notn_sr_val = 1
                    
                    if 'ROSCTL' in new_lic_type:
                        exim_code_val = 'RS'
                        exim_notn_val = 'ROSCTL'
                        exim_notn_sr_val = 1
                    elif 'RODTEP' in new_lic_type or 'RD' in new_lic_type:
                        exim_code_val = 'RD'
                        exim_notn_val = 'RODTEP'
                        exim_notn_sr_val = 1
                    elif 'DFIA' in new_lic_type:
                        exim_code_val = 'DF'
                        exim_notn_val = 'DFIA'
                        exim_notn_sr_val = 1
                    elif 'ADVANCE' in new_lic_type or 'ADV' in new_lic_type or 'AA' in new_lic_type:
                        exim_code_val = 'AA'
                        exim_notn_val = 'ADVANCE AUTHORISATION'
                        exim_notn_sr_val = 1
                    elif 'EPCG' in new_lic_type or 'EP' in new_lic_type:
                        exim_code_val = 'EP'
                        exim_notn_val = 'EPCG'
                        exim_notn_sr_val = 1
                        
                    if w_exim_code_idx != -1:
                        ws_write_items.cell(row=row_idx, column=w_exim_code_idx, value=exim_code_val)
                    if w_exim_notn_idx != -1:
                        ws_write_items.cell(row=row_idx, column=w_exim_notn_idx, value=exim_notn_val)
                    if w_exim_notn_sr_idx != -1:
                        ws_write_items.cell(row=row_idx, column=w_exim_notn_sr_idx, value=exim_notn_sr_val)
                    if w_duty_exemption_idx != -1:
                        ws_write_items.cell(row=row_idx, column=w_duty_exemption_idx, value="")
                elif k in affected_keys:
                    # Clear EXIM columns for failed/skipped items
                    if w_exim_code_idx != -1:
                        ws_write_items.cell(row=row_idx, column=w_exim_code_idx, value="")
                    if w_exim_notn_idx != -1:
                        ws_write_items.cell(row=row_idx, column=w_exim_notn_idx, value="")
                    if w_exim_notn_sr_idx != -1:
                        ws_write_items.cell(row=row_idx, column=w_exim_notn_sr_idx, value="")
                    if w_duty_exemption_idx != -1:
                        ws_write_items.cell(row=row_idx, column=w_duty_exemption_idx, value="")
            
            # Save rectified file
            dir_name = os.path.dirname(jd_path)
            orig_base = os.path.basename(jd_path)
            date_stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            ts_match = re.search(r'_(Processed|Rectified)_(\d{8}_\d{6})\.xlsx$', orig_base, re.IGNORECASE)
            if ts_match:
                output_name = orig_base[:ts_match.start()] + f"_Rectified_{date_stamp}.xlsx"
            else:
                output_name = re.sub(r'processed', 'Rectified', orig_base, flags=re.IGNORECASE)
                if 'rectified' not in output_name.lower():
                    clean_job_no_str = re.sub(r'[\\/*?:"<>|]', '_', job_no)
                    output_name = f"JobData_{clean_job_no_str}_Rectified_{date_stamp}.xlsx"
                    
            output_path = os.path.join(dir_name, output_name)
            
            wb_write.save(output_path)
            wb_write.close()
            
            # 2. Parse previously skipped items from existing Pending Items report if it exists
            previously_skipped = []
            orig_txt_path = jd_path.replace(".xlsx", "_Pending_Items_Low_Balance.txt")
            if os.path.exists(orig_txt_path):
                try:
                    with open(orig_txt_path, "r", encoding="utf-8") as f_prev:
                        lines = f_prev.readlines()
                    sep_idx = -1
                    for idx, line in enumerate(lines):
                        if "---" in line:
                            sep_idx = idx
                            break
                    if sep_idx != -1:
                        for line in lines[sep_idx+1:]:
                            line_str = line.strip()
                            if not line_str:
                                continue
                            parts = [p.strip() for p in line_str.split("|")]
                            if len(parts) >= 6:
                                duty_val = safe_float(parts[4].replace(",", ""))
                                previously_skipped.append({
                                    'Inv_Sr': parts[1],
                                    'Item_Sr': parts[2],
                                    'Invoice_No': parts[3],
                                    'Duty': duty_val,
                                    'Product_Desc': parts[5]
                                })
                            elif len(parts) >= 4:
                                duty_val = safe_float(parts[2].replace(",", ""))
                                previously_skipped.append({
                                    'Inv_Sr': "",
                                    'Item_Sr': "",
                                    'Invoice_No': parts[1],
                                    'Duty': duty_val,
                                    'Product_Desc': parts[3]
                                })
                except Exception as e:
                    self.safe_log(f"Warning: Could not parse previous pending items report: {e}")

            # Combine previously skipped and newly failed reallocation items
            all_pending = []
            seen_items = set()
            
            # Add previously skipped
            for item in previously_skipped:
                k = (item['Invoice_No'], item['Product_Desc'])
                if k not in seen_items:
                    seen_items.add(k)
                    all_pending.append({
                        'Inv_Sr': item['Inv_Sr'],
                        'Item_Sr': item['Item_Sr'],
                        'Invoice_No': item['Invoice_No'],
                        'Duty': item['Duty'],
                        'Product_Desc': item['Product_Desc']
                    })
            
            # Add newly failed reallocation items
            for item in skipped_reallocations:
                total_item_duty = sum(
                    safe_float(row[10])
                    for row in lic_rows_raw[1:]
                    if len(row) > 10 and str(row[0]).strip() == item['Inv_SrNo'] and str(row[1]).strip() == item['Item_SrNo']
                )
                k = (item['Invoice_No'], item['Product_Desc'])
                if k not in seen_items:
                    seen_items.add(k)
                    all_pending.append({
                        'Inv_Sr': item['Inv_SrNo'],
                        'Item_Sr': item['Item_SrNo'],
                        'Invoice_No': item['Invoice_No'],
                        'Duty': total_item_duty,
                        'Product_Desc': item['Product_Desc']
                    })

            if all_pending:
                skipped_txt_path = output_path.replace(".xlsx", "_Pending_Items_Low_Balance.txt")
                with open(skipped_txt_path, "w", encoding="utf-8") as f:
                    f.write(f"JOB NUMBER: {job_no}\n")
                    f.write("THE FOLLOWING ITEMS ARE PENDING DEBIT / REALLOCATION BECAUSE THE LICENSE BALANCE WAS TOO LOW TO COVER THEIR DUTY FULLY TO 0:\n\n")
                    f.write(f"{'No.':<4} | {'InvSrNo':<8} | {'ItemSrNo':<8} | {'Invoice No.':<20} | {'Duty (INR)':<15} | {'Product Description'}\n")
                    f.write("-" * 100 + "\n")
                    for idx, item in enumerate(all_pending, start=1):
                        f.write(f"{idx:<4} | {item['Inv_Sr']:<8} | {item['Item_Sr']:<8} | {item['Invoice_No']:<20} | {item['Duty']:<15,.2f} | {item['Product_Desc']}\n")
                self.safe_log(f"Saved {len(all_pending)} pending items list to: {os.path.basename(skipped_txt_path)}")
                
            self.safe_log("JobData file saved locally. Updating cloud database ledger...")
            
            # 2. Push to Google Sheets API
            url = self.google_sheet_url.get().strip()
            payload = {
                "action": "blacklist_license",
                "token": self.security_token.get().strip(),
                "lic_no": lics_to_blacklist,
                "job_no": job_no,
                "reason": reason,
                "reallocations": reallocations_api
            }
            
            response = requests.post(
                url, 
                json=payload, 
                headers={"Content-Type": "application/json", "Connection": "close"}, 
                timeout=30
            )
            if response.status_code != 200:
                raise ConnectionError(f"Cloud update failed with status code: {response.status_code}")
                
            try:
                res_bl = response.json()
            finally:
                response.close()
                
            if not res_bl.get("success"):
                raise ValueError(res_bl.get("error", "Unknown Error"))
                
            self.gui_queue.put(lambda: self._blacklist_reallocation_success(output_path))
            
        except Exception as e:
            err_msg = str(e)
            self.gui_queue.put(lambda: self._blacklist_reallocation_failure(err_msg))

    def _blacklist_reallocation_success(self, output_path):
        self.root.config(cursor="")
        self.blacklist_execute_btn.config(state='disabled')
        self.blacklist_parse_btn.config(state='normal')
        
        pending_txt_path = output_path.replace(".xlsx", "_Pending_Items_Low_Balance.txt")
        has_pending = os.path.exists(pending_txt_path)
        
        msg = f"Licenses blacklisted and debits re-allocated successfully!\n\nNew rectified file saved to:\n{os.path.basename(output_path)}"
        if has_pending:
            msg += f"\n\n⚠️ WARNING: Some items were skipped due to low license balances! A report listing all pending items has been saved to:\n{os.path.basename(pending_txt_path)}"
            
        messagebox.showinfo("Success", msg, parent=self.root)
        
        # Clear inputs
        self.blacklist_jd_path.set('')
        for item_id in self.blacklist_select_tree.get_children():
            self.blacklist_select_tree.delete(item_id)
        self.blacklist_selected_to_blacklist.clear()
        self.blacklist_reason_entry.delete(0, tk.END)
        self.blacklist_execute_btn.config(state='disabled')
        
        self.blacklist_summary_lbl.config(text="Please upload processed JobData and parse to view preview.")
        for item_id in self.blacklist_lic_tree.get_children():
            self.blacklist_lic_tree.delete(item_id)
        for item_id in self.blacklist_proposed_tree.get_children():
            self.blacklist_proposed_tree.delete(item_id)
            
        # Refresh blacklist and active databases
        self.refresh_blacklist()
        self.refresh_db_view()

    def _blacklist_reallocation_failure(self, error_msg):
        self.root.config(cursor="")
        self.blacklist_execute_btn.config(state='normal')
        self.blacklist_parse_btn.config(state='normal')
        messagebox.showerror("Error", f"Failed to re-allocate debits:\n\n{error_msg}", parent=self.root)

    def refresh_blacklist(self):
        url = self.google_sheet_url.get().strip()
        if not url:
            messagebox.showerror("Error", "Google Web App URL is not configured.")
            return
        self.root.config(cursor="watch")
        self.log("Loading blacklisted licenses from Google Sheets...")
        
        threading.Thread(target=self._refresh_blacklist_worker, args=(url,), daemon=True).start()

    def _refresh_blacklist_worker(self, url):
        try:
            payload = {
                "action": "fetch",
                "token": self.security_token.get().strip()
            }
            headers = {"Content-Type": "application/json", "Connection": "close"}
            response = requests.post(url, json=payload, headers=headers, timeout=20)
            if response.status_code != 200:
                raise ConnectionError(f"Cloud fetch failed with status: {response.status_code}")
                
            try:
                res = response.json()
            finally:
                response.close()
                
            if not res.get("success"):
                raise ValueError(res.get("error", "Unknown Error"))
                
            blacklist_rows = res.get("blacklist", [])
            self.gui_queue.put(lambda: self._refresh_blacklist_success(blacklist_rows))
        except Exception as e:
            err_msg = str(e)
            self.gui_queue.put(lambda: self._refresh_blacklist_failure(err_msg))

    def _refresh_blacklist_success(self, blacklist_rows):
        self.root.config(cursor="")
        for item_id in self.blacklist_tree.get_children():
            self.blacklist_tree.delete(item_id)
            
        self.blacklist_licenses = []
        if len(blacklist_rows) > 1:
            header = [str(h).strip().upper() for h in blacklist_rows[0]]
            lic_idx = header.index('LICNO/') if 'LICNO/' in header else 0
            date_idx = header.index('DATE BLACKLISTED') if 'DATE BLACKLISTED' in header else 1
            reason_idx = header.index('REASON') if 'REASON' in header else 2
            
            ldate_idx = header.index('LIC DATE') if 'LIC DATE' in header else -1
            port_idx = header.index('PORT OF REGISTRATION') if 'PORT OF REGISTRATION' in header else -1
            type_idx = header.index('LICENCE TYPE') if 'LICENCE TYPE' in header else -1
            val_idx = header.index('VALUE') if 'VALUE' in header else -1
            exp_idx = header.index('EXPIRY DATE') if 'EXPIRY DATE' in header else -1
            
            for row in blacklist_rows[1:]:
                if len(row) > lic_idx and row[lic_idx]:
                    lic_no = str(row[lic_idx]).strip()
                    date_bl = str(row[date_idx]).strip()
                    reason = str(row[reason_idx]).strip()
                    
                    self.blacklist_tree.insert('', 'end', values=(lic_no, date_bl, reason))
                    
                    self.blacklist_licenses.append({
                        'lic_no': lic_no,
                        'date_bl': date_bl,
                        'reason': reason,
                        'date': str(row[ldate_idx]).strip() if ldate_idx != -1 and ldate_idx < len(row) else "",
                        'port': str(row[port_idx]).strip() if port_idx != -1 and port_idx < len(row) else "",
                        'type': str(row[type_idx]).strip() if type_idx != -1 and type_idx < len(row) else "",
                        'val': str(row[val_idx]).strip() if val_idx != -1 and val_idx < len(row) else "",
                        'expiry': str(row[exp_idx]).strip() if exp_idx != -1 and exp_idx < len(row) else ""
                    })
        self.log(f"Successfully loaded {len(self.blacklist_licenses)} blacklisted licenses.")

    def _refresh_blacklist_failure(self, error_msg):
        self.root.config(cursor="")
        self.log(f"Failed to load blacklist: {error_msg}")
        messagebox.showerror("Error", f"Failed to refresh blacklist:\n\n{error_msg}", parent=self.root)

    def on_blacklist_row_selected(self, event):
        selected = self.blacklist_tree.selection()
        if not selected:
            self.rectify_restore_btn.config(state='disabled')
            return
            
        item_values = self.blacklist_tree.item(selected[0], 'values')
        lic_no = item_values[0]
        
        matched = None
        for l in self.blacklist_licenses:
            if l['lic_no'] == lic_no:
                matched = l
                break
                
        if matched:
            self.rectify_fields['lic_no'].set(matched['lic_no'])
            self.rectify_fields['date'].set(matched['date'])
            self.rectify_fields['port'].set(matched['port'])
            self.rectify_fields['type'].set(matched['type'])
            self.rectify_fields['val'].set(matched['val'])
            self.rectify_fields['expiry'].set(matched['expiry'])
            
            self.rectify_restore_btn.config(state='normal')

    def run_rectify_restore(self):
        lic_no = self.rectify_fields['lic_no'].get().strip()
        if not lic_no:
            return
            
        edited_data = {
            'date': self.rectify_fields['date'].get().strip(),
            'port': self.rectify_fields['port'].get().strip(),
            'type': self.rectify_fields['type'].get().strip(),
            'val': self.rectify_fields['val'].get().strip(),
            'expiry': self.rectify_fields['expiry'].get().strip()
        }
        
        if not (edited_data['date'] and edited_data['port'] and edited_data['type'] and edited_data['val'] and edited_data['expiry']):
            messagebox.showerror("Error", "Please fill in all details for rectification.", parent=self.root)
            return
            
        try:
            float(edited_data['val'])
        except ValueError:
            messagebox.showerror("Error", "License Value must be a valid number.", parent=self.root)
            return
            
        self.root.config(cursor="watch")
        self.rectify_restore_btn.config(state='disabled')
        self.log(f"Rectifying and restoring license {lic_no} to active pool...")
        
        threading.Thread(target=self._rectify_restore_worker, args=(lic_no, edited_data), daemon=True).start()

    def _rectify_restore_worker(self, lic_no, edited_data):
        try:
            url = self.google_sheet_url.get().strip()
            payload = {
                "action": "rectify_license",
                "token": self.security_token.get().strip(),
                "lic_no": lic_no,
                "edited_data": edited_data
            }
            headers = {"Content-Type": "application/json", "Connection": "close"}
            response = requests.post(url, json=payload, headers=headers, timeout=20)
            if response.status_code != 200:
                raise ConnectionError(f"Cloud update failed with status: {response.status_code}")
                
            try:
                res = response.json()
            finally:
                response.close()
                
            if not res.get("success"):
                raise ValueError(res.get("error", "Unknown Error"))
                
            self.gui_queue.put(lambda: self._rectify_restore_success())
        except Exception as e:
            err_msg = str(e)
            self.gui_queue.put(lambda: self._rectify_restore_failure(err_msg))

    def _rectify_restore_success(self):
        self.root.config(cursor="")
        messagebox.showinfo("Success", "License rectified and restored to the active pool successfully!", parent=self.root)
        for var in self.rectify_fields.values():
            var.set('')
        self.rectify_restore_btn.config(state='disabled')
        self.refresh_blacklist()
        self.refresh_db_view()

    def _rectify_restore_failure(self, error_msg):
        self.root.config(cursor="")
        self.rectify_restore_btn.config(state='normal')
        messagebox.showerror("Error", f"Failed to restore license:\n\n{error_msg}", parent=self.root)

    def _rectify_reallocation_success(self, output_path):
        self.root.config(cursor="")
        self.rectify_execute_btn.config(state='normal')
        self.rectify_parse_btn.config(state='normal')
        
        messagebox.showinfo("Success", f"License blacklisted and debits re-allocated successfully!\n\nNew rectified file saved to:\n{os.path.basename(output_path)}", parent=self.root)
        
        # Clear inputs
        self.rectify_jd_path.set('')
        self.rectify_lic_combo['values'] = []
        self.rectify_lic_combo.set('')
        self.rectify_reason_entry.delete(0, tk.END)
        self.rectify_execute_btn.config(state='disabled')
        
        # Refresh blacklist list and active list
        self.refresh_blacklist()
        self.refresh_db_view()

    def _rectify_reallocation_failure(self, error_msg):
        self.root.config(cursor="")
        self.rectify_execute_btn.config(state='normal')
        self.rectify_parse_btn.config(state='normal')
        messagebox.showerror("Error", f"Failed to re-allocate debits:\n\n{error_msg}", parent=self.root)

    def process_gui_queue(self):
        """Consume callbacks in the main thread to update the GUI safely."""
        try:
            while True:
                callback = self.gui_queue.get_nowait()
                try:
                    callback()
                except Exception as e:
                    self.log(f"GUI Callback Error: {e}")
        except queue.Empty:
            pass
        self.root.after(100, self.process_gui_queue)

    def log(self, msg: str):
        """Append a message to the logging pane."""
        self.log_txt.insert(tk.END, f"[{datetime.now().strftime('%H:%M:%S')}] {msg}\n")
        self.log_txt.see(tk.END)

    def safe_log(self, msg: str):
        """Append a message to the logging pane safely from any thread."""
        self.gui_queue.put(lambda: self.log(msg))

    def save_app_config(self):
        url = self.google_sheet_url.get().strip()
        token = self.security_token.get().strip()
        self.config["google_sheet_url"] = url
        self.config["security_token"] = token
        save_config(self.config)
        messagebox.showinfo("Success", "Configuration saved successfully!")
        if url:
            self.refresh_db_view()

    def browse_job_data(self):
        filename = filedialog.askopenfilename(title="Select JobData Excel File", filetypes=[("Excel Files", "*.xlsx")])
        if filename:
            self.job_data_path.set(filename)

    def browse_item_report(self):
        filename = filedialog.askopenfilename(title="Select Import Item Report Excel", filetypes=[("Excel Files", "*.xlsx")])
        if filename:
            self.item_report_path.set(filename)

    def browse_new_lic_excel(self):
        filename = filedialog.askopenfilename(title="Select Excel File with New Licenses", filetypes=[("Excel Files", "*.xlsx")])
        if filename:
            self.new_lic_excel_path.set(filename)

    def parse_sheet_date(self, date_str):
        if not date_str:
            return None
        date_str = str(date_str).split()[0].strip()
        # Supported formats including 2-digit years
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d", "%d-%b-%Y", "%d-%b-%y", "%d-%m-%y", "%m/%d/%y", "%y-%m-%d"):
            try:
                return datetime.strptime(date_str, fmt)
            except ValueError:
                pass
        try:
            return pd.to_datetime(date_str).to_pydatetime()
        except Exception:
            return None

    def format_input_date(self, date_str):
        if not date_str:
            return ""
        parsed = self.parse_sheet_date(date_str)
        if parsed:
            return parsed.strftime("%Y-%m-%d")
        return str(date_str).strip()

    def refresh_db_view(self):
        url = self.google_sheet_url.get().strip()
        if not url:
            return
        self.log("Refreshing database view from Google Sheet...")
        threading.Thread(target=self._refresh_db_view_worker, args=(url,), daemon=True).start()

    def _refresh_db_view_worker(self, url):
        try:
            payload = {
                "action": "fetch",
                "token": self.security_token.get().strip()
            }
            headers = {
                "Content-Type": "application/json",
                "Connection": "close"
            }
            response = requests.post(url, json=payload, headers=headers, timeout=20)
            if response.status_code != 200:
                raise ConnectionError(f"HTTP status {response.status_code}")
                
            try:
                res = response.json()
            except Exception:
                raise ValueError(humanize_error(response.text))
            finally:
                response.close()
                
            if not res.get("success"):
                raise ValueError(res.get("error", "Unknown Error"))
                
            data_rows = res.get("data", [])
            used_rows = res.get("used", [])
            self.gui_queue.put(lambda: self._refresh_db_view_success(data_rows, used_rows))
        except Exception as e:
            self.safe_log(f"Failed to refresh DB view: {e}")

    def _refresh_db_view_success(self, data_rows, used_rows):
        for item_id in self.db_tree.get_children():
            self.db_tree.delete(item_id)
            
        if not data_rows or len(data_rows) < 2:
            return
            
        header = [str(h).strip().upper() for h in data_rows[0]]
        
        try:
            lic_idx = header.index('LICNO/')
            date_idx = header.index('LIC DATE')
            port_idx = header.index('PORT OF REGISTRATION')
            type_idx = header.index('LICENCE TYPE')
            val_idx = header.index('VALUE')
            exp_idx = header.index('EXPIRY DATE')
            noboe_idx = header.index('COUNT OF JOB')
            used_idx = header.index('USED VALUE')
            bal_idx = header.index('BALANCE')
            job_idx = header.index('JOB NO')
        except ValueError as e:
            self.log(f"Database sheet format error: missing column: {e}")
            return
            
        self.active_licenses = []
        for r in data_rows[1:]:
            if len(r) <= max(lic_idx, bal_idx):
                continue
            
            base_val = safe_float(r[val_idx])
            used_val = safe_float(r[used_idx])
            bal_val = safe_float(r[bal_idx])
            
            lic_no = str(r[lic_idx]).split('.')[0].strip()
            
            exp_val = r[exp_idx]
            if isinstance(exp_val, str) and exp_val.strip() and exp_val.strip().lower() != 'nan':
                exp_date = self.parse_sheet_date(exp_val)
                exp_str = exp_date.strftime('%d-%b-%Y') if exp_date else exp_val
            else:
                exp_date = None
                exp_str = "N/A"
                
            lic_date_val = r[date_idx]
            if isinstance(lic_date_val, str) and lic_date_val.strip() and lic_date_val.strip().lower() != 'nan':
                lic_date = self.parse_sheet_date(lic_date_val)
                lic_date_str = lic_date.strftime('%d-%b-%Y') if lic_date else lic_date_val
            else:
                lic_date = None
                lic_date_str = "N/A"
                
            self.db_tree.insert('', 'end', values=(
                lic_no,
                lic_date_str,
                str(r[port_idx] or "").strip(),
                str(r[type_idx] or "").strip(),
                f"{base_val:.2f}",
                exp_str,
                str(r[noboe_idx] or 0),
                f"{used_val:.2f}",
                f"{bal_val:.2f}",
                str(r[job_idx] or "").strip()
            ))
            
            # Keep in-memory active licenses list populated
            self.active_licenses.append({
                'lic_no': lic_no,
                'port': str(r[port_idx] or "").strip(),
                'type': str(r[type_idx] or "").strip(),
                'val': base_val,
                'bal': bal_val,
                'expiry': exp_date,
                'reg_date': lic_date
            })
            
        # Parse and cache Used License sheet rows
        self.used_licenses = []
        if used_rows and len(used_rows) > 1:
            u_header = [str(h).strip().upper() for h in used_rows[0]]
            try:
                u_lic_idx = u_header.index('LICNO/')
                u_port_idx = u_header.index('PORT OF REGISTRATION')
                u_type_idx = u_header.index('LICENCE TYPE')
                u_val_idx = u_header.index('VALUE')
                u_bal_idx = u_header.index('BALANCE')
                u_exp_idx = u_header.index('EXPIRY DATE')
                u_date_idx = u_header.index('LIC DATE')
                
                for r in used_rows[1:]:
                    if len(r) <= max(u_lic_idx, u_bal_idx):
                        continue
                    u_lic_no = str(r[u_lic_idx]).split('.')[0].strip()
                    
                    u_exp_val = r[u_exp_idx]
                    if isinstance(u_exp_val, str) and u_exp_val.strip() and u_exp_val.strip().lower() != 'nan':
                        u_exp_date = self.parse_sheet_date(u_exp_val)
                    else:
                        u_exp_date = None
                        
                    u_lic_date_val = r[u_date_idx]
                    if isinstance(u_lic_date_val, str) and u_lic_date_val.strip() and u_lic_date_val.strip().lower() != 'nan':
                        u_lic_date = self.parse_sheet_date(u_lic_date_val)
                    else:
                        u_lic_date = None
                        
                    self.used_licenses.append({
                        'lic_no': u_lic_no,
                        'port': str(r[u_port_idx] or "").strip(),
                        'type': str(r[u_type_idx] or "").strip(),
                        'val': safe_float(r[u_val_idx]),
                        'bal': safe_float(r[u_bal_idx]),
                        'expiry': u_exp_date,
                        'reg_date': u_lic_date
                    })
            except Exception as e:
                self.log(f"Failed parsing used licenses: {e}")
                
        self.log("Database view updated.")

    def parse_pasted_text(self, text_content):
        licenses = []
        lines = text_content.strip().split('\n')
        for line in lines:
            if not line.strip():
                continue
            if '\t' in line:
                parts = [p.strip() for p in line.split('\t')]
            else:
                parts = [p.strip() for p in re.split(r'\s{2,}', line)]
                
            if len(parts) >= 5:
                lic_no = parts[0].strip()
                date_str = self.format_input_date(parts[1].strip())
                port = parts[2].strip()
                lic_type = parts[3].strip()
                val_str = parts[4].strip()
                
                raw_exp = parts[5].strip() if len(parts) > 5 else ""
                if not raw_exp or raw_exp.lower() == 'nan':
                    parsed_date = self.parse_sheet_date(date_str)
                    if parsed_date:
                        exp_date = parsed_date + timedelta(days=720)
                        expiry_str = exp_date.strftime("%Y-%m-%d")
                    else:
                        expiry_str = ""
                else:
                    expiry_str = self.format_input_date(raw_exp)
                
                noboe_val = int(safe_float(parts[6].strip())) if len(parts) > 6 and parts[6].strip() else 0
                used_val = safe_float(parts[7].strip()) if len(parts) > 7 and parts[7].strip() else 0.0
                bal_val = safe_float(parts[8].strip()) if len(parts) > 8 and parts[8].strip() else safe_float(val_str)
                job_no_val = parts[9].strip() if len(parts) > 9 else ""
                
                if lic_no and date_str and val_str:
                    licenses.append({
                        "lic_no": lic_no,
                        "date": date_str,
                        "port": port,
                        "type": lic_type,
                        "val": safe_float(val_str),
                        "expiry": expiry_str,
                        "noboe": noboe_val,
                        "used_val": used_val,
                        "balance": bal_val,
                        "job_no": job_no_val
                    })
        return licenses

    def parse_excel_licenses(self, file_path):
        # Prioritize visible sheet named "Sheet1", fallback to first visible sheet
        visible_sheet = 0
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            found_sheet1 = False
            for name in wb.sheetnames:
                if name.strip().upper() == "SHEET1" and wb[name].sheet_state == 'visible':
                    visible_sheet = name
                    found_sheet1 = True
                    break
            if not found_sheet1:
                for name in wb.sheetnames:
                    if wb[name].sheet_state == 'visible':
                        visible_sheet = name
                        break
            wb.close()
        except Exception:
            pass
            
        df = pd.read_excel(file_path, sheet_name=visible_sheet, engine='openpyxl')
        cols = [str(c).strip().upper() for c in df.columns]
        
        lic_col = -1
        date_col = -1
        port_col = -1
        type_col = -1
        val_col = -1
        exp_col = -1
        noboe_col = -1
        used_col = -1
        bal_col = -1
        job_col = -1
        
        for idx, col in enumerate(cols):
            col_str = str(col).strip().upper()
            if 'LICNO' in col_str or 'LIC REF' in col_str or 'LICENSE NO' in col_str:
                lic_col = idx
            elif 'EXPIRY' in col_str:
                exp_col = idx
            elif 'LIC DATE' in col_str or ('DATE' in col_str and 'EXPIRY' not in col_str):
                date_col = idx
            elif 'PORT' in col_str:
                port_col = idx
            elif 'TYPE' in col_str:
                type_col = idx
            elif 'USED' in col_str:
                used_col = idx
            elif 'BAL' in col_str:
                bal_col = idx
            elif 'COUNT' in col_str or 'NO OF BOE' in col_str or 'BOE' in col_str:
                noboe_col = idx
            elif 'JOB' in col_str:
                job_col = idx
            elif 'VALUE' in col_str or 'VAL' in col_str:
                val_col = idx
                
        if lic_col == -1: lic_col = 0
        if date_col == -1: date_col = 1
        if port_col == -1: port_col = 2
        if type_col == -1: type_col = 3
        if val_col == -1: val_col = 4
        if exp_col == -1 and len(df.columns) > 5: exp_col = 5
        
        licenses = []
        for idx, row in df.iterrows():
            row_list = list(row)
            if len(row_list) > max(lic_col, val_col) and not pd.isna(row_list[lic_col]):
                lic_no = str(row_list[lic_col]).replace(".0", "").strip()
                if not lic_no:
                    continue
                # Skip non-numeric rows like "Legend" headers
                if not lic_no.isdigit():
                    continue
                
                raw_date = row_list[date_col]
                if pd.isna(raw_date):
                    date_str = ""
                elif isinstance(raw_date, datetime):
                    date_str = raw_date.strftime("%Y-%m-%d")
                else:
                    date_str = self.format_input_date(str(raw_date).split()[0])
                    
                raw_exp = row_list[exp_col] if exp_col != -1 and exp_col < len(row_list) else None
                if pd.isna(raw_exp) or not str(raw_exp).strip():
                    # Calculate LIC DATE + 720 days
                    parsed_date = self.parse_sheet_date(date_str)
                    if parsed_date:
                        exp_date = parsed_date + timedelta(days=720)
                        exp_str = exp_date.strftime("%Y-%m-%d")
                    else:
                        exp_str = ""
                elif isinstance(raw_exp, datetime):
                    exp_str = raw_exp.strftime("%Y-%m-%d")
                else:
                    exp_str = self.format_input_date(str(raw_exp).split()[0])
                
                # Parse additional historical columns if present in Excel
                noboe_val = 0
                if noboe_col != -1 and noboe_col < len(row_list):
                    val_noboe = row_list[noboe_col]
                    noboe_val = int(safe_float(val_noboe)) if not pd.isna(val_noboe) else 0
                    
                used_val = 0.0
                if used_col != -1 and used_col < len(row_list):
                    val_used = row_list[used_col]
                    used_val = safe_float(val_used) if not pd.isna(val_used) else 0.0
                    
                bal_val = safe_float(row_list[val_col])
                if bal_col != -1 and bal_col < len(row_list):
                    val_bal = row_list[bal_col]
                    bal_val = safe_float(val_bal) if not pd.isna(val_bal) else bal_val
                    
                job_no_val = ""
                if job_col != -1 and job_col < len(row_list):
                    val_job = row_list[job_col]
                    job_no_val = str(val_job).strip() if not pd.isna(val_job) else ""
                    if job_no_val.lower() == "nan":
                        job_no_val = ""
                
                licenses.append({
                    "lic_no": lic_no,
                    "date": date_str,
                    "port": str(row_list[port_col]).strip() if not pd.isna(row_list[port_col]) else "",
                    "type": str(row_list[type_col]).strip() if not pd.isna(row_list[type_col]) else "",
                    "val": safe_float(row_list[val_col]),
                    "expiry": exp_str,
                    "noboe": noboe_val,
                    "used_val": used_val,
                    "balance": bal_val,
                    "job_no": job_no_val
                })
        return licenses

    def push_pasted_licenses(self):
        text = self.paste_txt.get("1.0", tk.END).strip()
        if not text:
            messagebox.showerror("Error", "Please paste license data first.")
            return
            
        url = self.google_sheet_url.get().strip()
        if not url:
            messagebox.showerror("Error", "Google Web App URL is not configured.")
            return
            
        try:
            licenses = self.parse_pasted_text(text)
            if not licenses:
                messagebox.showerror("Error", "Could not parse any valid license rows. Please verify formatting.")
                return
                
            self.paste_btn.config(state='disabled')
            self.excel_btn.config(state='disabled')
            self.root.config(cursor="watch")
            
            self.log(f"Pasting {len(licenses)} licenses to Google Sheets...")
            threading.Thread(target=self._push_licenses_worker, args=(url, licenses, "paste"), daemon=True).start()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to parse pasted data: {e}")

    def push_excel_licenses(self):
        path = self.new_lic_excel_path.get().strip()
        if not path or not os.path.exists(path):
            messagebox.showerror("Error", "Please select a valid Excel file.")
            return
            
        url = self.google_sheet_url.get().strip()
        if not url:
            messagebox.showerror("Error", "Google Web App URL is not configured.")
            return
            
        try:
            licenses = self.parse_excel_licenses(path)
            if not licenses:
                messagebox.showerror("Error", "No license data found in the Excel sheet.")
                return
                
            self.paste_btn.config(state='disabled')
            self.excel_btn.config(state='disabled')
            self.root.config(cursor="watch")
            
            self.log(f"Uploading {len(licenses)} licenses from Excel to Google Sheets...")
            threading.Thread(target=self._push_licenses_worker, args=(url, licenses, "excel"), daemon=True).start()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to parse Excel file: {e}")

    def _push_licenses_worker(self, url, licenses_payload, source_type):
        try:
            to_push = licenses_payload
            self.safe_log(f"Pushing {len(to_push)} licenses to Google Sheets cloud in bulk...")
            
            payload = {
                "action": "add_licenses",
                "token": self.security_token.get().strip(),
                "licenses": to_push
            }
            headers = {
                "Content-Type": "application/json",
                "Connection": "close"
            }
            
            try:
                response = requests.post(url, json=payload, headers=headers, timeout=30)
                if response.status_code != 200:
                    raise ConnectionError(f"HTTP status {response.status_code}")
                
                try:
                    res = response.json()
                except Exception:
                    raise ValueError(humanize_error(response.text))
                finally:
                    response.close()
                
                if not res.get("success"):
                    raise ValueError(res.get("error", "Unknown Error"))
                    
                added_count = res.get("added", len(to_push))
                skipped_count = res.get("skipped", 0)
                
                self.safe_log(f"Pusher finished: {added_count} added successfully, {skipped_count} duplicates skipped.")
                self.gui_queue.put(lambda: self._push_licenses_success(source_type, added_count, skipped_count, []))
                
            except Exception as e:
                self.safe_log(f"Bulk license push failed: {e}")
                self.gui_queue.put(lambda: self._push_licenses_success(source_type, 0, len(to_push), [(l['lic_no'], str(e)) for l in to_push]))
                
        except Exception as e:
            err_msg = str(e)
            self.safe_log(f"Critical error during license push: {err_msg}")
            self.gui_queue.put(lambda: self._push_licenses_failure(err_msg))

    def _push_licenses_success(self, source_type, success_count, skipped_count, error_details):
        self.paste_btn.config(state='normal')
        self.excel_btn.config(state='normal')
        self.root.config(cursor="")
        
        if source_type == "paste":
            self.paste_txt.delete("1.0", tk.END)
        else:
            self.new_lic_excel_path.set("")
            
        summary_msg = f"Upload process completed:\n\n"
        summary_msg += f"✅ Successfully added: {success_count} new license(s)\n"
        if skipped_count > 0:
            summary_msg += f"ℹ️ Skipped (duplicates already in master): {skipped_count} license(s)\n"
        if error_details:
            summary_msg += f"❌ Skipped/Failed (cloud validation error): {len(error_details)} license(s)\n\n"
            summary_msg += "Validation details:\n"
            for lic_no, err in error_details[:5]:
                clean_err = str(err)
                if "Google Sheets Script Error" in clean_err:
                    clean_err = clean_err.split("\n")[0]
                summary_msg += f" - Lic {lic_no}: {clean_err}\n"
            if len(error_details) > 5:
                summary_msg += f" - ... and {len(error_details) - 5} more."
                
        messagebox.showinfo("Upload Summary", summary_msg)
        self.refresh_db_view()

    def _push_licenses_failure(self, error_msg):
        self.paste_btn.config(state='normal')
        self.excel_btn.config(state='normal')
        self.root.config(cursor="")
        messagebox.showerror("Error", f"Failed to push licenses to cloud: {error_msg}")

    def on_tree_click(self, event):
        """Toggle checkbox status on treeview row click anywhere on the row."""
        item_id = self.lic_tree.identify_row(event.y)
        if not item_id:
            return
            
        values = list(self.lic_tree.item(item_id, 'values'))
        if values[0] == "☑":
            values[0] = "☐"
        else:
            values[0] = "☑"
            
        current_tags = list(self.lic_tree.item(item_id, 'tags'))
        if values[0] == "☑":
            if 'checked' not in current_tags:
                current_tags.append('checked')
        else:
            if 'checked' in current_tags:
                current_tags.remove('checked')
                
        self.lic_tree.item(item_id, values=values, tags=tuple(current_tags))
        self.recalc_selected_duty_metrics()

    def recalc_selected_duty_metrics(self):
        """Calculate selected license sum and compare with required duty."""
        self.selected_duty_covered = 0.0
        selected_lics = []
        for item_id in self.lic_tree.get_children():
            values = self.lic_tree.item(item_id, 'values')
            if values[0] == "☑":
                try:
                    bal_val = safe_float(values[4])
                    self.selected_duty_covered += bal_val
                    selected_lics.append({
                        'lic_no': values[1],
                        'bal': bal_val,
                        'type': values[3]
                    })
                except ValueError:
                    pass
        
        self.estimated_debit, _ = self.simulate_debit(selected_lics)
        self.update_summary_card()

    def simulate_debit(self, selected_lics):
        """Simulate greedy allocation in strict database (top-to-bottom) order.
        
        For each item:
          1. Walk the license list top-to-bottom.
          2. The FIRST license with available balance determines the type for that item.
          3. Continue consuming from subsequent licenses of the SAME type until the
             item's duty is fully covered (all-or-nothing per type).
          4. If the chosen type cannot cover the item fully, try the next type that
             the NEXT available license belongs to.
        """
        if not hasattr(self, 'item_duties') or not self.item_duties:
            return 0.0, {}
            
        lics = [dict(l) for l in selected_lics]
        license_debit_totals = {lic['lic_no']: 0.0 for lic in lics}
        
        for item_data in self.item_duties:
            duty = item_data['duty']
            if duty <= 0.0:
                continue
            
            # Walk top-to-bottom to find the first available license
            # Its type becomes the candidate type for this item
            tried_types = set()
            covered = False
            
            for start_idx, start_lic in enumerate(lics):
                lic_type = str(start_lic['type']).strip().upper()
                if lic_type in tried_types:
                    continue
                max_start = start_lic['bal'] - 1.00
                if max_start <= 0.01:
                    continue
                    
                # Found a candidate type — check if ALL licenses of this type
                # (in order) can cover the full duty
                total_avail = sum(
                    max(0.0, l['bal'] - 1.00)
                    for l in lics
                    if str(l['type']).strip().upper() == lic_type
                )
                
                if total_avail < duty:
                    tried_types.add(lic_type)
                    continue
                
                # Allocate from licenses of this type in top-to-bottom order
                duty_remaining = duty
                for lic in lics:
                    if duty_remaining <= 0.005:
                        break
                    if str(lic['type']).strip().upper() != lic_type:
                        continue
                    max_debitable = lic['bal'] - 1.00
                    if max_debitable <= 0.01:
                        continue
                        
                    if max_debitable >= duty_remaining:
                        debit_amt = duty_remaining
                        lic['bal'] -= debit_amt
                        license_debit_totals[lic['lic_no']] += debit_amt
                        duty_remaining = 0.0
                    else:
                        debit_amt = max_debitable
                        lic['bal'] = 1.00
                        license_debit_totals[lic['lic_no']] += debit_amt
                        duty_remaining -= debit_amt
                covered = True
                break  # Item covered, move to next item
                    
        total_estimated_debit = sum(license_debit_totals.values())
        return total_estimated_debit, license_debit_totals

    def load_and_analyze_data(self):
        """Parse files, extract job summary, load licenses, and update UI."""
        jd_path = self.job_data_path.get()
        ir_path = self.item_report_path.get()
        
        if not (jd_path and ir_path):
            messagebox.showerror("Error", "Please configure both JobData and Item Report Excel files first.")
            return
            
        if not (os.path.exists(jd_path) and os.path.exists(ir_path)):
            messagebox.showerror("Error", "One or more files do not exist at the specified paths.")
            return

        self.root.config(cursor="watch")
        self.load_btn.config(state='disabled')
        self.run_btn.config(state='disabled')
        self.log("Starting analysis of files in background thread...")
        
        scheme_dropdown_val = "Auto-Detect"
        
        threading.Thread(
            target=self._load_and_analyze_worker,
            args=(jd_path, ir_path, scheme_dropdown_val),
            daemon=True
        ).start()

    def _load_and_analyze_worker(self, jd_path, ir_path, scheme_dropdown_val):
        try:
            self.safe_log("Pre-validating Item Report and JobData files...")
            
            # 1. Ultra-fast pre-check: read Job No directly from xlsx zip (no pandas/openpyxl)
            report_job_no = _quick_extract_job_no(ir_path)
            if report_job_no is None:
                raise ValueError("The selected file is not a valid Item Report (missing 'Job No' column or data).")
                
            # 2. Extract job number from JobData filename
            jd_filename = os.path.basename(jd_path)
            jd_clean = jd_filename
            if jd_clean.lower().startswith("jobdata_"):
                jd_clean = jd_clean[8:]
            jd_clean = re.sub(r'_\d{8}_\d{6}\.xlsx$', '', jd_clean, flags=re.IGNORECASE)
            jd_clean = re.sub(r'\.xlsx$', '', jd_clean, flags=re.IGNORECASE)
            
            # Normalize both
            norm_report = re.sub(r'[^a-zA-Z0-9]', '', report_job_no).upper()
            norm_filename = re.sub(r'[^a-zA-Z0-9]', '', jd_clean).upper()
            
            if norm_report != norm_filename:
                raise ValueError(
                    f"File Mismatch Error!\n\n"
                    f"The selected JobData Excel file does not match the loaded Item Report.\n\n"
                    f"• Item Report Job No: {report_job_no}\n"
                    f"• JobData Excel File: {jd_filename}\n\n"
                    f"Please load matching files for the same Job."
                )
                
            self.safe_log("Reading Import Item Report...")
            df_rep = pd.read_excel(ir_path, engine='openpyxl')
            
            cols = df_rep.columns.tolist()
            req_cols = ['BE No', 'BE Date', 'Job No', 'Assessable Value (INR)', 'Basic Duty Rate', 'Exim Scheme Code', 'Quantity', 'Unit', 'Product Desc']
            missing = [r for r in req_cols if r not in cols]
            if missing:
                raise ValueError(f"Import Item Report is missing required column(s): {missing}")
            
            self.safe_log("Reading JobData sheets (GENERAL, INVOICES, ITEMS) for analysis...")
            wb_jd = openpyxl.load_workbook(jd_path, read_only=True)
            if 'GENERAL' not in wb_jd.sheetnames:
                raise ValueError("GENERAL sheet not found in JobData workbook.")
            ws_gen = wb_jd['GENERAL']
            rows_gen = list(ws_gen.iter_rows(values_only=True))
            headers_gen = [str(h).strip().upper() for h in rows_gen[0]]
            
            if 'CUSTOMSHOUSECODE' not in headers_gen:
                raise ValueError("Column CUSTOMSHOUSECODE not found in JobData GENERAL sheet.")
            port_col_idx = headers_gen.index('CUSTOMSHOUSECODE')
            import_port = str(rows_gen[1][port_col_idx]).strip()
            
            # Map invoices to get InvSrNo
            if 'INVOICES' not in wb_jd.sheetnames:
                raise ValueError("INVOICES sheet not found in JobData workbook.")
            ws_inv = wb_jd['INVOICES']
            inv_rows = list(ws_inv.iter_rows(values_only=True))
            inv_header = [str(h).strip() for h in inv_rows[0]]
            inv_sr_idx = inv_header.index('InvSrNo')
            inv_no_idx = inv_header.index('Invoice_No')
            
            inv_map = {}
            for r in inv_rows[1:]:
                if r[inv_sr_idx] is not None and r[inv_no_idx] is not None:
                    inv_map[str(r[inv_no_idx]).strip()] = str(r[inv_sr_idx]).strip()
            
            # Read ITEMS sheet
            if 'ITEMS' not in wb_jd.sheetnames:
                raise ValueError("ITEMS sheet not found in JobData workbook.")
            ws_items = wb_jd['ITEMS']
            items_rows = list(ws_items.iter_rows(values_only=True))
            items_header = [str(h).strip() for h in items_rows[0]]
            item_inv_sr_idx = items_header.index('InvSrNo')
            item_sr_idx = items_header.index('ItemSrNo')
            desc_idx = items_header.index('Product_Description')
            qty_idx = items_header.index('QTY')
            cth_idx = items_header.index('CTH')
            exim_code_idx = items_header.index('Exim_Code') if 'Exim_Code' in items_header else -1
            
            wb_jd.close()
            
            # Index items_rows by InvSrNo for O(1) subset lookup
            items_by_inv_sr = {}
            for r_idx, item_row in enumerate(items_rows[1:], start=2):
                if item_row[item_inv_sr_idx] is not None:
                    itm_inv_sr = str(item_row[item_inv_sr_idx]).strip()
                    if itm_inv_sr not in items_by_inv_sr:
                        items_by_inv_sr[itm_inv_sr] = []
                    items_by_inv_sr[itm_inv_sr].append((r_idx, item_row))
            
            # Match each item from the report to check for Exim_Code 14 and store scheme
            item_duties = []
            restricted_count = 0
            matched_rows = set()
            
            for idx, r_row in df_rep.iterrows():
                rep_inv_no = str(r_row['Invoice No']).strip()
                rep_desc = str(r_row['Product Desc']).strip()
                rep_qty = safe_float(r_row['Quantity'])
                rep_cth = str(r_row['CTH']).strip()
                rep_scheme = str(r_row['Exim Scheme Code']).strip().upper() if not pd.isna(r_row['Exim Scheme Code']) else "RD"
                if rep_scheme in ('NAN', ''):
                    rep_scheme = "RD"
                
                av = safe_float(r_row['Assessable Value (INR)'])
                rate = safe_float(r_row['Basic Duty Rate'])
                
                inv_sr = inv_map.get(rep_inv_no)
                is_restricted = False
                
                if inv_sr and exim_code_idx != -1:
                    candidates = items_by_inv_sr.get(inv_sr, [])
                    for r_idx, item_row in candidates:
                        if r_idx in matched_rows:
                            continue
                        itm_desc = str(item_row[desc_idx]).strip() if item_row[desc_idx] is not None else ""
                        itm_qty = safe_float(item_row[qty_idx])
                        itm_cth = str(item_row[cth_idx]).strip() if item_row[cth_idx] is not None else ""
                        
                        if itm_desc == rep_desc and abs(itm_qty - rep_qty) < 0.01 and itm_cth == rep_cth:
                            matched_rows.add(r_idx)
                            exim_code_val = str(item_row[exim_code_idx]).strip()
                            if exim_code_val == '14' or exim_code_val == '14.0':
                                is_restricted = True
                            break
                
                if is_restricted:
                    duty = 0.0
                    restricted_count += 1
                    self.safe_log(f"Row {idx+1}: Item '{rep_desc[:25]}' has Exim_Code 14 (Restricted). Skipping license allocation.")
                else:
                    duty = round(av * rate / 100.0, 2)
                    
                item_duties.append({'duty': duty, 'scheme': rep_scheme})
            
            required_duty = round(sum(d['duty'] for d in item_duties), 2)
            
            raw_be_no = df_rep.loc[0, 'BE No']
            be_no = str(raw_be_no).strip() if not pd.isna(raw_be_no) else "nan"
            
            be_date = df_rep.loc[0, 'BE Date']
            if pd.isna(be_date):
                be_date_str = "nan"
            elif isinstance(be_date, datetime):
                be_date_str = be_date.strftime('%d-%b-%Y')
            else:
                be_date_str = str(be_date).split()[0]
                
            job_no = str(df_rep.loc[0, 'Job No']).strip()
            
            raw_scheme_code = df_rep.loc[0, 'Exim Scheme Code']
            if scheme_dropdown_val == "Auto-Detect":
                if pd.isna(raw_scheme_code) or str(raw_scheme_code).strip().lower() in ('nan', ''):
                    scheme_code = "RD"
                    self.safe_log("Exim Scheme Code is empty in file. Defaulting to RD (RODTEP).")
                else:
                    scheme_code = str(raw_scheme_code).strip()
            else:
                if scheme_dropdown_val == "RODTEP":
                    scheme_code = "RD"
                else:
                    scheme_code = scheme_dropdown_val
                self.safe_log(f"Using manual License Scheme override: {scheme_dropdown_val}")
                
            total_items = len(df_rep)
            
            job_info = {
                'be_no': be_no,
                'be_date': be_date,
                'be_date_str': be_date_str,
                'job_no': job_no,
                'scheme': scheme_code,
                'total_items': total_items,
                'import_port': import_port
            }
            
            self.safe_log(f"Job Details: Job No: {job_no}, BE No: {be_no}, Import Port: {import_port}, Scheme: {scheme_code}")
            if restricted_count > 0:
                self.safe_log(f"Note: Found {restricted_count} restricted item(s) (Exim_Code 14) which won't use licenses.")
            self.safe_log(f"Total Checklist Items: {total_items}, Required License Duty: {required_duty:,.2f} INR")
            
            # Load active licenses from Google Sheet API
            url = self.google_sheet_url.get().strip()
            if not url:
                raise ValueError("Google Web App URL is not configured. Please set it in the Database Master tab.")
                
            self.safe_log("Loading active licenses from Google Sheets cloud...")
            headers = {
                "Content-Type": "application/json",
                "Connection": "close"
            }
            payload = {
                "action": "fetch",
                "token": self.security_token.get().strip()
            }
            response = requests.post(url, json=payload, headers=headers, timeout=30)
            if response.status_code != 200:
                raise ConnectionError(f"Cloud request failed with status: {response.status_code}")
                
            try:
                res_data = response.json()
            except Exception:
                raise ValueError(humanize_error(response.text))
            finally:
                response.close()
                
            if not res_data.get("success"):
                raise ValueError(f"Cloud DB returned error: {res_data.get('error', 'Unknown Error')}")
                
            # Extract already debited job numbers from the cloud ledger
            debited_jobs = set()
            boe_rows = res_data.get("boe", [])
            if len(boe_rows) > 1:
                header_boe = [str(h).strip().upper() for h in boe_rows[0]]
                if 'LIC USED IN THE JOB' in header_boe:
                    job_idx_boe = header_boe.index('LIC USED IN THE JOB')
                    for row in boe_rows[1:]:
                        if len(row) > job_idx_boe and row[job_idx_boe]:
                            debited_jobs.add(str(row[job_idx_boe]).strip().upper())
            
            if report_job_no.upper() in debited_jobs:
                raise ValueError(
                    f"Duplicate Debit Protection!\n\n"
                    f"Job '{report_job_no}' has ALREADY been debited in the cloud database ledger.\n\n"
                    f"License allocation for this job is blocked to prevent duplicate debiting."
                )
                
            scrip_rows = res_data.get("data", [])
            if not scrip_rows or len(scrip_rows) < 2:
                self.safe_log("Warning: No license records found in the cloud database.")
                scrip_rows = []
                
            header_scrip = [str(h).strip().upper() for h in scrip_rows[0]]
            
            lic_idx = header_scrip.index('LICNO/')
            bal_idx = header_scrip.index('BALANCE')
            port_idx = header_scrip.index('PORT OF REGISTRATION')
            type_idx = header_scrip.index('LICENCE TYPE')
            exp_idx = header_scrip.index('EXPIRY DATE')
            val_idx = header_scrip.index('VALUE')
            date_idx = header_scrip.index('LIC DATE')
            
            active_licenses = []
            for idx, r in enumerate(scrip_rows[1:], start=2):
                if len(r) <= max(lic_idx, bal_idx, port_idx, type_idx, exp_idx) or r[lic_idx] is None:
                    continue
                lic_type = str(r[type_idx]).strip().upper() if r[type_idx] is not None else ""
                
                bal = r[bal_idx]
                if bal is not None:
                    try:
                        bal_f = safe_float(bal)
                        if bal_f > 0.00:
                            exp_val = r[exp_idx]
                            exp_date = self.parse_sheet_date(exp_val)
                            reg_date_val = r[date_idx]
                            reg_date = self.parse_sheet_date(reg_date_val)
                            
                            active_licenses.append({
                                'lic_no': str(r[lic_idx]).split('.')[0].strip(),
                                'port': str(r[port_idx]).strip(),
                                'type': str(r[type_idx]).strip(),
                                'val': safe_float(r[val_idx]),
                                'bal': bal_f,
                                'expiry': exp_date,
                                'reg_date': reg_date
                            })
                    except ValueError:
                        pass
            
            self.safe_log(f"Successfully loaded {len(active_licenses)} active licenses.")
            self.gui_queue.put(lambda: self._load_and_analyze_success(job_info, active_licenses, required_duty, item_duties))
            
        except Exception as e:
            err_msg = str(e)
            self.gui_queue.put(lambda: self._load_and_analyze_failure(err_msg))

    def _load_and_analyze_success(self, job_info, active_licenses, required_duty, item_duties):
        self.job_info = job_info
        self.active_licenses = active_licenses
        self.required_duty = required_duty
        self.item_duties = item_duties
        
        self.populate_license_table()
        self.recalc_selected_duty_metrics()
        
        self.root.config(cursor="")
        self.load_btn.config(state='normal')
        self.run_btn.config(state='normal')
        self.log("Analysis completed successfully.")

    def _load_and_analyze_failure(self, error_msg):
        self.root.config(cursor="")
        self.load_btn.config(state='normal')
        self.run_btn.config(state='disabled')
        self.log(f"Error during analysis: {error_msg}")
        
        # Clear inputs specifically for Mismatch and Duplicate Debit errors
        is_mismatch = "File Mismatch Error" in error_msg
        is_duplicate_debit = "Duplicate Debit Protection" in error_msg
        if is_mismatch or is_duplicate_debit:
            self.job_data_path.set("")
            self.item_report_path.set("")
            # Reset summary panel to default placeholder
            for widget in self.summary_text_frame.winfo_children():
                widget.destroy()
            ttk.Label(self.summary_text_frame, text="Please load the JobData and Item Report files to view summary.", style='Summary.TLabel').pack(anchor='w', pady=10)
            self.log("JobData and Item Report paths cleared due to verification error.")
            
        messagebox.showerror("Error", f"Failed to analyze files:\n\n{error_msg}", parent=self.root)

    def populate_license_table(self):
        """Fill treeview table and auto-check the licenses that will actually be utilized, retaining database order."""
        for item_id in self.lic_tree.get_children():
            self.lic_tree.delete(item_id)
            
        # Run simulation to see which licenses are actually used
        candidate_lics = [item for item in self.active_licenses if item['bal'] >= 3.00]
        _, debit_totals = self.simulate_debit(candidate_lics)
        
        for item in self.active_licenses:
            is_used = (debit_totals.get(item['lic_no'], 0.0) > 0.0)
            item['auto_selected'] = is_used
            
            exp_str = item['expiry'].strftime('%d-%b-%Y') if item['expiry'] is not None else "N/A"
            selected_str = "☑" if item['auto_selected'] else "☐"
            
            is_near = False
            if item['expiry'] is not None:
                days_to_expiry = (item['expiry'] - datetime.now()).days
                if days_to_expiry <= 30 and item['val'] > 500:
                    is_near = True
            
            tags = []
            if selected_str == "☑":
                tags.append('checked')
            if is_near:
                tags.append('near_expiry')
                
            self.lic_tree.insert('', 'end', values=(
                selected_str,
                item['lic_no'],
                item['port'],
                item['type'],
                f"{item['bal']:.2f}",
                exp_str
            ), tags=tuple(tags))

    def update_summary_card(self):
        """Update left summary panel with calculated details and shortfall indicators."""
        for widget in self.summary_text_frame.winfo_children():
            widget.destroy()
            
        if not self.job_info:
            return
            
        grid_params = {'anchor': 'w', 'pady': 3}
        
        be_no_val = self.job_info['be_no'] if str(self.job_info['be_no']).lower() != 'nan' else "Pending / Empty"
        be_date_val = self.job_info['be_date_str'] if str(self.job_info['be_date_str']).lower() != 'nan' else "Pending / Empty"
        
        ttk.Label(self.summary_text_frame, text=f"Job Number: {self.job_info['job_no']}", style='Summary.TLabel').pack(**grid_params)
        ttk.Label(self.summary_text_frame, text=f"Bill of Entry No: {be_no_val}", style='Summary.TLabel').pack(**grid_params)
        ttk.Label(self.summary_text_frame, text=f"Bill of Entry Date: {be_date_val}", style='Summary.TLabel').pack(**grid_params)
        ttk.Label(self.summary_text_frame, text=f"Total Items in Checklist: {self.job_info['total_items']}", style='Summary.TLabel').pack(**grid_params)
        
        separator = ttk.Separator(self.summary_text_frame, orient='horizontal')
        separator.pack(fill='x', pady=10)
        
        cash_payment = max(0.0, self.required_duty - self.estimated_debit)
        
        ttk.Label(self.summary_text_frame, text=f"Total Duty to Pay: {self.required_duty:,.2f} INR", font=(FONT_FAMILY, 10, 'bold')).pack(**grid_params)
        ttk.Label(self.summary_text_frame, text=f"Selected License Capacity: {self.selected_duty_covered:,.2f} INR", font=(FONT_FAMILY, 10, 'bold')).pack(**grid_params)
        ttk.Label(self.summary_text_frame, text=f"Duty Covered by License (Exemption): {self.estimated_debit:,.2f} INR", font=(FONT_FAMILY, 10, 'bold')).pack(**grid_params)
        ttk.Label(self.summary_text_frame, text=f"Remaining Duty to Pay in Cash: {cash_payment:,.2f} INR", font=(FONT_FAMILY, 10, 'bold')).pack(**grid_params)
        
        if cash_payment <= 0.02:
            status_text = "Status: Fully Covered (No cash payment required)"
            status_color = COLOR_ACCENT
        else:
            status_text = f"Status: Partially Covered (Cash payment of {cash_payment:,.2f} INR required)"
            status_color = COLOR_ERROR
            
        ttk.Label(self.summary_text_frame, text=status_text, foreground=status_color, font=(FONT_FAMILY, 11, 'bold')).pack(**grid_params)

    def run_license_automation(self):
        """Execute license allocation, modify JobData, modify central database, and notify user."""
        jd_path = self.job_data_path.get()
        ir_path = self.item_report_path.get()
        
        selected_lic_nos = []
        for item_id in self.lic_tree.get_children():
            values = self.lic_tree.item(item_id, 'values')
            if values[0] == "☑":
                selected_lic_nos.append(values[1])
                
        if not selected_lic_nos:
            messagebox.showerror("Error", "No licenses selected. Please check at least one license.")
            return
            
        if self.estimated_debit < self.required_duty - 0.02:
            cash_needed = self.required_duty - self.estimated_debit
            ans = messagebox.askyesno("Warning", f"The selected licenses do not cover the total required duty. You will need to pay {cash_needed:,.2f} INR in cash to customs. Do you want to proceed?")
            if not ans:
                return
  
        self.root.config(cursor="watch")
        self.load_btn.config(state='disabled')
        self.run_btn.config(state='disabled')
        self.log("Initiating license debiting process in background thread...")
        
        threading.Thread(
            target=self._run_license_automation_worker,
            args=(jd_path, ir_path, selected_lic_nos),
            daemon=True
        ).start()

    def _run_license_automation_worker(self, jd_path, ir_path, selected_lic_nos):
        try:
            selected_lics = []
            for lic_no in selected_lic_nos:
                for item in self.active_licenses:
                    if item['lic_no'] == lic_no:
                        selected_lics.append(dict(item))
                        break
            
            self.safe_log("Reading Import Item Report...")
            df_rep = pd.read_excel(ir_path, engine='openpyxl')
            
            self.safe_log("Opening JobData Excel for writing (preserving formulas)...")
            wb_jd = openpyxl.load_workbook(jd_path, data_only=False)
            ws_inv = wb_jd['INVOICES']
            inv_rows = list(ws_inv.iter_rows(values_only=True))
            inv_header = [str(h).strip() for h in inv_rows[0]]
            inv_sr_idx = inv_header.index('InvSrNo')
            inv_no_idx = inv_header.index('Invoice_No')
            
            inv_map = {}
            for r in inv_rows[1:]:
                if r[inv_sr_idx] is not None and r[inv_no_idx] is not None:
                    inv_map[str(r[inv_no_idx]).strip()] = str(r[inv_sr_idx]).strip()
            
            ws_items = wb_jd['ITEMS']
            items_rows = list(ws_items.iter_rows(values_only=True))
            items_header = [str(h).strip() for h in items_rows[0]]
            item_inv_sr_idx = items_header.index('InvSrNo')
            item_sr_idx = items_header.index('ItemSrNo')
            desc_idx = items_header.index('Product_Description')
            qty_idx = items_header.index('QTY')
            cth_idx = items_header.index('CTH')
            
            exim_code_idx = items_header.index('Exim_Code') if 'Exim_Code' in items_header else -1
            exim_notn_idx = items_header.index('Exim_Notn') if 'Exim_Notn' in items_header else -1
            exim_notn_sr_idx = items_header.index('Exim_NotnSrNo') if 'Exim_NotnSrNo' in items_header else -1
            duty_exemption_idx = items_header.index('DUTY_ExemptionType') if 'DUTY_ExemptionType' in items_header else -1
            
            self.safe_log("Running allocation algorithm...")
            job_license_rows = []
            skipped_items = []
            license_debit_totals = {lic['lic_no']: 0.0 for lic in selected_lics}
            
            # Index items_rows by InvSrNo for O(1) subset lookup
            items_by_inv_sr = {}
            for r_idx, item_row in enumerate(items_rows[1:], start=2):
                if item_row[item_inv_sr_idx] is not None:
                    itm_inv_sr = str(item_row[item_inv_sr_idx]).strip()
                    if itm_inv_sr not in items_by_inv_sr:
                        items_by_inv_sr[itm_inv_sr] = []
                    items_by_inv_sr[itm_inv_sr].append((r_idx, item_row))
            
            matched_rows = set()
            
            for idx, r_row in df_rep.iterrows():
                rep_inv_no = str(r_row['Invoice No']).strip()
                rep_desc = str(r_row['Product Desc']).strip()
                rep_qty = safe_float(r_row['Quantity'])
                rep_cth = str(r_row['CTH']).strip()
                
                av = safe_float(r_row['Assessable Value (INR)'])
                rate = safe_float(r_row['Basic Duty Rate'])
                
                inv_sr = inv_map.get(rep_inv_no)
                item_sr = None
                matched_row_idx = None
                is_restricted = False
                
                if inv_sr:
                    candidates = items_by_inv_sr.get(inv_sr, [])
                    for r_idx, item_row in candidates:
                        if r_idx in matched_rows:
                            continue
                        itm_desc = str(item_row[desc_idx]).strip() if item_row[desc_idx] is not None else ""
                        itm_qty = safe_float(item_row[qty_idx])
                        itm_cth = str(item_row[cth_idx]).strip() if item_row[cth_idx] is not None else ""
                        
                        if itm_desc == rep_desc and abs(itm_qty - rep_qty) < 0.01 and itm_cth == rep_cth:
                            item_sr = str(item_row[item_sr_idx]).strip()
                            matched_row_idx = r_idx
                            matched_rows.add(r_idx)
                            if exim_code_idx != -1:
                                exim_code_val = str(item_row[exim_code_idx]).strip()
                                if exim_code_val == '14' or exim_code_val == '14.0':
                                    is_restricted = True
                            break
                
                if not item_sr:
                    self.safe_log(f"Warning: Could not match item row {idx} (Desc: {rep_desc[:30]}...) to ITEMS sheet. Skipping.")
                    continue
                
                if is_restricted:
                    self.safe_log(f"Skipping item {idx+1} (Desc: {rep_desc[:25]}) - Restricted item (Exim_Code 14). Row in ITEMS sheet is kept as it is.")
                    continue
                
                duty = round(av * rate / 100.0, 2)
                
                if duty == 0.0:
                    matching_lic = selected_lics[0] if len(selected_lics) > 0 else None
                    if matching_lic:
                        job_license_rows.append({
                            'Inv_SrNo': inv_sr,
                            'Item_SrNo': item_sr,
                            'Invoice_No': rep_inv_no,
                            'License_No': matching_lic['lic_no'],
                            'License_Date': matching_lic['reg_date'].strftime('%d-%b-%Y') if matching_lic['reg_date'] else "",
                            'License_RegNo': matching_lic['lic_no'],
                            'License_RegDate': matching_lic['reg_date'].strftime('%d-%b-%Y') if matching_lic['reg_date'] else "",
                            'Reg_Port': matching_lic['port'],
                            'CIF_Value': av,
                            'DebitDeutyValue': 0.00,
                            'DebitQuantity': rep_qty,
                            'DebitQuantityUnitCode': r_row['Unit'],
                            'Product_Desc': rep_desc,
                            'Assessable_Value': av,
                            'Basic_Duty_Rate': rate,
                            'Basic_Duty_Value': 0.00
                        })
                        # Update exim values based on license type
                        lic_type = str(matching_lic['type']).strip().upper()
                        if 'ROSCTL' in lic_type:
                            exim_code_val = 'RS'
                            exim_notn_val = 'ROSCTL'
                            exim_notn_sr_val = 1
                        elif 'RODTEP' in lic_type or 'RD' in lic_type:
                            exim_code_val = 'RD'
                            exim_notn_val = 'RODTEP'
                            exim_notn_sr_val = 1
                        elif 'DFIA' in lic_type:
                            exim_code_val = 'DF'
                            exim_notn_val = 'DFIA'
                            exim_notn_sr_val = 1
                        elif 'EPCG' in lic_type:
                            exim_code_val = 'EP'
                            exim_notn_val = 'EPCG'
                            exim_notn_sr_val = 1
                        elif 'ADVANCE' in lic_type:
                            exim_code_val = 'AA'
                            exim_notn_val = 'ADVANCE AUTHORISATION'
                            exim_notn_sr_val = 1
                        else:
                            exim_code_val = 'RD'
                            exim_notn_val = 'RODTEP'
                            exim_notn_sr_val = 1
                            
                        if matched_row_idx:
                            if exim_code_idx != -1:
                                ws_items.cell(row=matched_row_idx, column=exim_code_idx + 1, value=exim_code_val)
                            if exim_notn_idx != -1:
                                ws_items.cell(row=matched_row_idx, column=exim_notn_idx + 1, value=exim_notn_val)
                            if exim_notn_sr_idx != -1:
                                ws_items.cell(row=matched_row_idx, column=exim_notn_sr_idx + 1, value=exim_notn_sr_val)
                            if duty_exemption_idx != -1:
                                ws_items.cell(row=matched_row_idx, column=duty_exemption_idx + 1, value="")
                    continue

                # Find which type can cover this item by walking licenses top-to-bottom
                allocated_type = None
                tried_types = set()
                for lic in selected_lics:
                    lic_type = str(lic['type']).strip().upper()
                    if lic_type in tried_types:
                        continue
                    if lic['bal'] - 1.00 <= 0.01:
                        continue
                    # This is the first license in line with balance — check if its type can cover fully
                    total_avail = sum(
                        max(0.0, l['bal'] - 1.00)
                        for l in selected_lics
                        if str(l['type']).strip().upper() == lic_type
                    )
                    if total_avail >= duty:
                        allocated_type = lic_type
                        break
                    else:
                        tried_types.add(lic_type)
                        
                if not allocated_type:
                    self.safe_log(f"Skipping item {idx+1} (Duty: {duty:.2f} INR) - Insufficient remaining balance in any single license type to cover it fully to 0.")
                    skipped_items.append({
                        'index': idx + 1,
                        'invoice_no': rep_inv_no,
                        'desc': rep_desc,
                        'qty': rep_qty,
                        'cth': rep_cth,
                        'duty': duty,
                        'inv_sr': inv_sr,
                        'item_sr': item_sr
                    })
                    continue

                duty_remaining = duty
                last_used_lic = None
                for lic in selected_lics:
                    if duty_remaining <= 0.005:
                        break
                    if str(lic['type']).strip().upper() != allocated_type:
                        continue
                    max_debitable = lic['bal'] - 1.00
                    if max_debitable <= 0.01:
                        continue
                        
                    last_used_lic = lic
                    if max_debitable >= duty_remaining:
                        debit_amt = duty_remaining
                        lic['bal'] -= debit_amt
                        ratio = debit_amt / duty
                        job_license_rows.append({
                            'Inv_SrNo': inv_sr,
                            'Item_SrNo': item_sr,
                            'Invoice_No': rep_inv_no,
                            'License_No': lic['lic_no'],
                            'License_Date': lic['reg_date'].strftime('%d-%b-%Y') if lic['reg_date'] else "",
                            'License_RegNo': lic['lic_no'],
                            'License_RegDate': lic['reg_date'].strftime('%d-%b-%Y') if lic['reg_date'] else "",
                            'Reg_Port': lic['port'],
                            'CIF_Value': av,
                            'DebitDeutyValue': round(debit_amt, 2),
                            'DebitQuantity': round(rep_qty * ratio, 3),
                            'DebitQuantityUnitCode': r_row['Unit'],
                            'Product_Desc': rep_desc,
                            'Assessable_Value': av,
                            'Basic_Duty_Rate': rate,
                            'Basic_Duty_Value': duty
                        })
                        license_debit_totals[lic['lic_no']] += debit_amt
                        duty_remaining = 0
                    else:
                        debit_amt = max_debitable
                        lic['bal'] = 1.00
                        ratio = debit_amt / duty
                        job_license_rows.append({
                            'Inv_SrNo': inv_sr,
                            'Item_SrNo': item_sr,
                            'Invoice_No': rep_inv_no,
                            'License_No': lic['lic_no'],
                            'License_Date': lic['reg_date'].strftime('%d-%b-%Y') if lic['reg_date'] else "",
                            'License_RegNo': lic['lic_no'],
                            'License_RegDate': lic['reg_date'].strftime('%d-%b-%Y') if lic['reg_date'] else "",
                            'Reg_Port': lic['port'],
                            'CIF_Value': av,
                            'DebitDeutyValue': round(debit_amt, 2),
                            'DebitQuantity': round(rep_qty * ratio, 3),
                            'DebitQuantityUnitCode': r_row['Unit'],
                            'Product_Desc': rep_desc,
                            'Assessable_Value': av,
                            'Basic_Duty_Rate': rate,
                            'Basic_Duty_Value': duty
                        })
                        license_debit_totals[lic['lic_no']] += debit_amt
                        duty_remaining -= debit_amt
                        
                # After the debiting loop for this item has processed (and if it was successfully debited),
                # we update the EXIM columns in ws_items!
                if duty_remaining == 0 and last_used_lic and matched_row_idx:
                    lic_type = str(last_used_lic['type']).strip().upper()
                    if 'ROSCTL' in lic_type:
                        exim_code_val = 'RS'
                        exim_notn_val = 'ROSCTL'
                        exim_notn_sr_val = 1
                    elif 'RODTEP' in lic_type or 'RD' in lic_type:
                        exim_code_val = 'RD'
                        exim_notn_val = 'RODTEP'
                        exim_notn_sr_val = 1
                    elif 'DFIA' in lic_type:
                        exim_code_val = 'DF'
                        exim_notn_val = 'DFIA'
                        exim_notn_sr_val = 1
                    elif 'EPCG' in lic_type:
                        exim_code_val = 'EP'
                        exim_notn_val = 'EPCG'
                        exim_notn_sr_val = 1
                    elif 'ADVANCE' in lic_type:
                        exim_code_val = 'AA'
                        exim_notn_val = 'ADVANCE AUTHORISATION'
                        exim_notn_sr_val = 1
                    else:
                        exim_code_val = 'RD'
                        exim_notn_val = 'RODTEP'
                        exim_notn_sr_val = 1
                        
                    if exim_code_idx != -1:
                        ws_items.cell(row=matched_row_idx, column=exim_code_idx + 1, value=exim_code_val)
                    if exim_notn_idx != -1:
                        ws_items.cell(row=matched_row_idx, column=exim_notn_idx + 1, value=exim_notn_val)
                    if exim_notn_sr_idx != -1:
                        ws_items.cell(row=matched_row_idx, column=exim_notn_sr_idx + 1, value=exim_notn_sr_val)
                    if duty_exemption_idx != -1:
                        ws_items.cell(row=matched_row_idx, column=duty_exemption_idx + 1, value="")

            self.safe_log("Writing to JobData LICENSE sheet...")
            if 'LICENSE' not in wb_jd.sheetnames:
                ws_lic = wb_jd.create_sheet('LICENSE')
            else:
                ws_lic = wb_jd['LICENSE']
                
            ws_lic.delete_rows(2, ws_lic.max_row)
            
            headers = [
                'Inv_SrNo', 'Item_SrNo', 'License_RefNo', 'License_No', 'License_Date',
                'License_RegNo', 'License_RegDate', 'Reg_Port', 'License_ItemSrNo',
                'CIF_Value', 'DebitDeutyValue', 'DebitQuantity', 'DebitQuantityUnitCode'
            ]
            for col_num, header in enumerate(headers, 1):
                ws_lic.cell(row=1, column=col_num, value=header)
                
            for row_idx, data in enumerate(job_license_rows, start=2):
                ws_lic.cell(row=row_idx, column=1, value=data['Inv_SrNo'])
                ws_lic.cell(row=row_idx, column=2, value=data['Item_SrNo'])
                ws_lic.cell(row=row_idx, column=3, value=None)
                ws_lic.cell(row=row_idx, column=4, value=data['License_No'])
                ws_lic.cell(row=row_idx, column=5, value=data['License_Date'])
                ws_lic.cell(row=row_idx, column=6, value=data['License_RegNo'])
                ws_lic.cell(row=row_idx, column=7, value=data['License_RegDate'])
                ws_lic.cell(row=row_idx, column=8, value=data['Reg_Port'])
                ws_lic.cell(row=row_idx, column=9, value=1)
                ws_lic.cell(row=row_idx, column=10, value=data['CIF_Value'])
                ws_lic.cell(row=row_idx, column=11, value=data['DebitDeutyValue'])
                ws_lic.cell(row=row_idx, column=12, value=data['DebitQuantity'])
                ws_lic.cell(row=row_idx, column=13, value=data['DebitQuantityUnitCode'])
            
            dir_name = os.path.dirname(jd_path)
            date_stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            clean_job_no_str = re.sub(r'[\\/*?:"<>|]', '_', self.job_info['job_no'])
            job_name = f"JobData_{clean_job_no_str}_Processed_{date_stamp}.xlsx"
            output_jd_path = os.path.join(dir_name, job_name)
            
            self.safe_log("Saving processed JobData Excel...")
            wb_jd.save(output_jd_path)
            wb_jd.close()
            self.safe_log(f"JobData saved successfully to: {os.path.basename(output_jd_path)}")
            
            if skipped_items:
                skipped_txt_path = output_jd_path.replace(".xlsx", "_Pending_Items_Low_Balance.txt")
                with open(skipped_txt_path, "w", encoding="utf-8") as f:
                    f.write(f"JOB NUMBER: {self.job_info.get('job_no', 'Unknown')}\n")
                    f.write("THE FOLLOWING ITEMS ARE PENDING DEBIT / REALLOCATION BECAUSE THE LICENSE BALANCE WAS TOO LOW TO COVER THEIR DUTY FULLY TO 0:\n\n")
                    f.write(f"{'No.':<4} | {'InvSrNo':<8} | {'ItemSrNo':<8} | {'Invoice No.':<20} | {'Duty (INR)':<15} | {'Product Description'}\n")
                    f.write("-" * 100 + "\n")
                    for i, item in enumerate(skipped_items, start=1):
                        f.write(f"{i:<4} | {item.get('inv_sr', ''):<8} | {item.get('item_sr', ''):<8} | {item['invoice_no']:<20} | {item['duty']:<15,.2f} | {item['desc']}\n")
                self.safe_log(f"Saved {len(skipped_items)} pending items (low balance) list to: {os.path.basename(skipped_txt_path)}")

            # 6. Push Debits to Google Sheets cloud
            self.safe_log("Preparing debit data to push to Google Sheets cloud...")
            debits_payload = []
            
            for data in job_license_rows:
                debits_payload.append({
                    "lic_no": data['License_No'],
                    "val": data['DebitDeutyValue'],
                    "job_no": self.job_info['job_no'],
                    "inv_no": data['Invoice_No'],
                    "inv_sr_no": data['Inv_SrNo'],
                    "item_sr_no": data['Item_SrNo'],
                    "desc": data['Product_Desc'],
                    "av": data['CIF_Value'],
                    "rate": data['Basic_Duty_Rate'],
                    "duty_val": data['Basic_Duty_Value']
                })
                
            self.safe_log("Pushing debits to Google Sheets cloud...")
            url = self.google_sheet_url.get().strip()
            payload = {
                "action": "debit",
                "token": self.security_token.get().strip(),
                "debits": debits_payload
            }
            headers = {
                "Content-Type": "application/json",
                "Connection": "close"
            }
            response = requests.post(url, json=payload, headers=headers, timeout=30)
            if response.status_code != 200:
                raise ConnectionError(f"Cloud update request failed with status: {response.status_code}")
                
            try:
                res_data = response.json()
            except Exception:
                raise ValueError(humanize_error(response.text))
            finally:
                response.close()
                
            if not res_data.get("success"):
                raise ValueError(res_data.get("error", "Unknown Error"))
                
            self.safe_log("Google Sheets updated successfully.")
            self.gui_queue.put(lambda: self._run_license_automation_success(output_jd_path))
            
        except Exception as e:
            err_msg = str(e)
            self.gui_queue.put(lambda: self._run_license_automation_failure(err_msg))

    def _run_license_automation_success(self, output_jd_path):
        self.root.config(cursor="")
        self.load_btn.config(state='normal')
        self.run_btn.config(state='disabled')
        
        pending_txt_path = output_jd_path.replace(".xlsx", "_Pending_Items_Low_Balance.txt")
        has_pending = os.path.exists(pending_txt_path)
        
        msg = f"License automation finished successfully!\n\n1. JobData updated and saved to:\n{os.path.basename(output_jd_path)}\n\n2. Google Sheets updated successfully in cloud."
        if has_pending:
            msg += f"\n\n⚠️ WARNING: Some items were skipped due to low license balances! A report listing all pending items has been saved to:\n{os.path.basename(pending_txt_path)}"
            
        messagebox.showinfo("Success", msg)
        
        # Clear input file paths after successful processing
        self.job_data_path.set("")
        self.item_report_path.set("")
        self.job_info = None
        self.item_duties = []
        self.required_duty = 0.0
        self.selected_duty_covered = 0.0
        self.estimated_debit = 0.0
        
        # Automatically refresh database view in the background to get updated balances
        self.refresh_db_view()
        
        # Clear treeview and summary
        for item_id in self.lic_tree.get_children():
            self.lic_tree.delete(item_id)
        for widget in self.summary_text_frame.winfo_children():
            widget.destroy()
        ttk.Label(self.summary_text_frame, text="Please load the JobData and Item Report files to view summary.", style='Summary.TLabel').pack(anchor='w', pady=10)
        
        self.log("All inputs cleared. Ready for next job.")

    def _run_license_automation_failure(self, error_msg):
        self.root.config(cursor="")
        self.load_btn.config(state='normal')
        self.run_btn.config(state='normal')
        self.log(f"Error executing license allocation: {error_msg}")
        messagebox.showerror("Error", f"Failed to complete license allocation:\n\n{error_msg}", parent=self.root)

if __name__ == "__main__":
    root = tk.Tk()
    app = LicenseAutomationApp(root)
    root.mainloop()
